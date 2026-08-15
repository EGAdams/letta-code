"""Non-destructive expense highlighting for supporting documents.

The stable application contracts live in ``document_annotation_contracts``.
This module implements format strategies and the service; the composition root
wires those strategies to IO-near adapters.
"""

from __future__ import annotations

import hashlib
import math
import os
import re
import statistics
import threading
from html import escape
from dataclasses import replace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Sequence

from codex_image_region_fallback import CodexCliImageRegionFallbackMatcher
from document_annotation_contracts import (
    AnnotationResult,
    ExpenseEvidence,
    IExpenseDocumentAnnotationService,
    IExpenseDocumentAnnotator,
    IExpenseReferenceResolver,
    IImageRegionFallbackMatcher,
    ImageRegionMatch,
)

ANNOTATION_SCHEMA_VERSION = 23


def _valid_image_region(
    region: tuple[float, float, float, float],
    image_width: int,
    image_height: int,
) -> bool:
    if len(region) != 4 or not all(
        isinstance(value, (int, float)) and math.isfinite(value)
        for value in region
    ):
        return False
    left, top, right, bottom = region
    return (
        0 <= left < right <= image_width
        and 0 <= top < bottom <= image_height
    )


_TOKEN_RE = re.compile(r"[a-z0-9]+")
_TOTAL_WORD_RE = re.compile(r"\btotal\b", re.I)
_LABELED_TOTAL_RE = re.compile(r"\btotal\b", re.I)

# Tesseract's output for the Diners Club annual-fee scan is stable but
# imperfect: the printed word ``ANNUAL`` is commonly returned as
# ``ADINUAL``/``ADINUALFEE``. Keep these corrections deliberately small and
# token-based. A broad edit-distance match here would make unrelated merchant
# names look like the expense being annotated.
_OCR_TOKEN_ALIASES = {
    "adinual": ("annual",),
    "adinualfee": ("annual", "fee"),
    "adnual": ("annual",),
    "anual": ("annual",),
    "inual": ("annual",),
}


def _token_list(value: str) -> list[str]:
    """Usable tokens in printed order — the payee comes before its city."""
    seen: list[str] = []
    for raw_token in _TOKEN_RE.findall(str(value or "").lower()):
        aliases = _OCR_TOKEN_ALIASES.get(raw_token, (raw_token,))
        for token in aliases:
            if len(token) > 2 and token not in {"the", "and", "for", "inc", "llc"}:
                if token not in seen:
                    seen.append(token)
    return seen


def _tokens(value: str) -> set[str]:
    return set(_token_list(value))


def _payee_head_tokens(description: str) -> set[str]:
    """The first tokens of a description — the payee's own name.

    A statement prints "PAYEE CITY STATE", so the trailing tokens are shared by
    every row from the same town ("GRAND RAPIDS MI") and identify nothing. Only
    the head is identity.
    """
    return set(_token_list(description)[:2])


def _amount_decimal(value: object) -> Decimal | None:
    raw = str(value or "").strip().replace("$", "").replace(",", "")
    raw = raw.strip("()")
    try:
        return abs(Decimal(raw))
    except (InvalidOperation, ValueError):
        return None


def _region_area(region: object) -> float:
    """Return a comparable geometric area for candidate precision ranking."""
    if (
        isinstance(region, tuple)
        and len(region) == 4
        and all(isinstance(value, (int, float)) for value in region)
    ):
        return max(0.0, float(region[2] - region[0])) * max(
            0.0, float(region[3] - region[1])
        )
    if (
        isinstance(region, tuple)
        and len(region) == 2
        and hasattr(region[1], "x0")
        and hasattr(region[1], "y0")
        and hasattr(region[1], "x1")
        and hasattr(region[1], "y1")
    ):
        return max(0.0, float(region[1].x1 - region[1].x0)) * max(
            0.0, float(region[1].y1 - region[1].y0)
        )
    return float("inf")


def _amount_matches(text: str, amount: Decimal | None) -> bool:
    if amount is None:
        return False
    number = f"{amount:.2f}"
    whole, cents = number.split(".")
    grouped = f"{int(whole):,}.{cents}"
    # Avoid matching 25.00 inside 125.00 or an account/reference number.
    if re.search(
        rf"(?<![\d.])(?:-\s*|\(\s*)?\$?\s*(?:{re.escape(number)}|"
        rf"{re.escape(grouped)})(?!\d)",
        text,
        re.I,
    ):
        return True

    # Card-statement scans sometimes lose the decimal point and retain the
    # amount as ``(9500 )``. Only accept that compact form at the end of the
    # OCR line; accepting it in the middle would confuse a reference number
    # with the charge amount.
    compact = f"{whole}{cents}"
    if re.search(
        rf"(?<!\d){re.escape(compact)}(?=\s*[)\]}}.,;:/\\-]*\s*$)",
        text,
        re.I,
    ):
        return True

    # A leading stroke is also read as ``3`` on this statement. Accept that
    # artifact only at the terminal amount position, and only as the known
    # compact/decimal forms; row identity still has to agree independently.
    return bool(
        re.search(
            rf"(?<!\d)3(?:{re.escape(compact)}|{re.escape(whole)}\.{re.escape(cents)})"
            rf"(?=\s*[)\]}}.,;:/\\-]*\s*$)",
            text,
            re.I,
        )
    )


def _foreign_amount_count(text: str, amount: Decimal | None) -> int:
    """Count decimal amounts in ``text`` other than the target amount.

    Evidence-aware OCR windows can accidentally join two neighboring card
    rows. Such a window may contain the target amount and still score well,
    but the unrelated second amount is strong evidence that it is not one
    transaction row. Repeated copies of the target amount are intentionally
    ignored because OCR commonly emits those from overlapping columns.
    """
    target = amount.quantize(Decimal("0.01")) if amount is not None else None
    target_count = 0
    count = 0
    for match in re.finditer(
        r"(?<![\d.])(?:-\s*|\(\s*)?\$?\s*(\d[\d,]*\.\d{2})(?!\d)",
        str(text or ""),
        re.I,
    ):
        try:
            value = Decimal(match.group(1).replace(",", ""))
        except InvalidOperation:
            continue
        if target is not None and value.quantize(Decimal("0.01")) == target:
            target_count += 1
        else:
            count += 1
    # A statement summary may legitimately repeat the same charge amount in
    # two adjacent labeled rows. Do not penalize that known same-amount shape;
    # the dangerous composite-window case has one target amount plus a
    # different neighboring amount.
    return count if target_count == 1 else 0


def _labeled_total_amount_stem_matches(
    text: str,
    amount: Decimal | None,
) -> bool:
    """Recover a TOTAL whose final cent digit alone was lost to handwriting."""
    if amount is None or not _LABELED_TOTAL_RE.search(text):
        return False
    whole, cents = f"{amount:.2f}".split(".")
    stem = f"{whole}.{cents[0]}"
    return bool(re.search(rf"(?<![\d.])\$?\s*{re.escape(stem)}(?!\d)", text, re.I))


def _date_variants(value: str) -> set[str]:
    raw = str(value or "").strip()
    parsed = None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m-%d-%Y", "%m-%d-%y"):
        try:
            parsed = datetime.strptime(raw, fmt).date()
            break
        except ValueError:
            continue
    if parsed is None:
        return {raw.lower()} if raw else set()
    return {
        parsed.strftime("%Y-%m-%d"),
        parsed.strftime("%m/%d/%Y"),
        parsed.strftime("%m/%d/%y"),
        parsed.strftime("%m/%d"),
        parsed.strftime("%m-%d-%Y"),
        parsed.strftime("%m-%d-%y"),
        parsed.strftime("%m-%d"),
        f"{parsed.month}/{parsed.day}/{parsed.year}",
        f"{parsed.month}/{parsed.day}/{str(parsed.year)[2:]}",
        f"{parsed.month}/{parsed.day}",
        f"{parsed.month}-{parsed.day}",
        f"{parsed.strftime('%B')} {parsed.day}, {parsed.year}",
        f"{parsed.strftime('%b')} {parsed.day}, {parsed.year}",
    }


def _date_matches(text: str, evidence_date: str) -> bool:
    lowered = text.lower()
    return any(
        value and value.lower() in lowered
        for value in _date_variants(evidence_date)
    )


def _complete_date_matches(text: str, evidence_date: str) -> bool:
    """Match a printed calendar date, never a bare M/D mentioned in prose."""
    variants = {
        value
        for value in _date_variants(evidence_date)
        if re.search(r"(?:^|[-/])\d{2,4}$", value)
        and value.count("/") + value.count("-") == 2
    }
    lowered = text.lower()
    return any(value and value.lower() in lowered for value in variants)


def _statement_row_date_matches(text: str, evidence_date: str) -> bool:
    """Match the bare M/D a credit-card statement row opens with.

    Card statements print the year once, in the billing-cycle header, and start
    every transaction row with ``MM/DD MM/DD`` (posting and transaction date).
    ``_complete_date_matches`` demands a year, so on those scans it can never
    fire and the identity-only path was unreachable — the Choice Privileges
    freezer scan of 2026-07-29 lost its box for exactly this reason. Anchoring
    at the start of the line keeps the looser M/D out of prose, where a
    "payment due by 03/21" sentence would otherwise read as a row date.
    """
    variants = {
        value
        for value in _date_variants(evidence_date)
        if value.count("/") + value.count("-") == 1
    }
    stripped = str(text or "").lstrip(" |[](){}<>*.,:;-_\"'`~")
    return any(
        value and stripped.lower().startswith(value.lower()) for value in variants
    )


def _has_leading_calendar_date(text: str) -> bool:
    """Whether OCR starts with a transaction-like M/D or M/D/Y date."""
    return bool(
        re.match(
            r"^\s*\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?(?=\D|$)",
            str(text or ""),
        )
    )


# A tie at or above this score is decisive; below it, two equally-ranked rows
# reject each other rather than boxing the wrong one.
_DECISIVE_SCORE = 10
# Ceiling for a match that never confirmed the amount.  Held below
# ``_DECISIVE_SCORE`` on purpose, so an identity-only win can never survive a
# tie with a second row that looks just as much like this expense.
_IDENTITY_ONLY_CEILING = 9


def _strong_description_score(
    description_tokens: set[str],
    description_hits: int,
) -> int:
    """Score a long, distinctive description when OCR loses date and amount."""
    if len(description_tokens) < 8 or description_hits < 5:
        return -1
    coverage = description_hits / len(description_tokens)
    if coverage < 0.5:
        return -1
    return min(
        _IDENTITY_ONLY_CEILING,
        5 + round(coverage * 4),
    )


def _description_only_score(
    description_tokens: set[str],
    description_hits: int,
) -> int:
    """Score a compact, distinctive full description without date/amount.

    Card-summary layouts can print the charge date in a separate header and
    split the label/card identifier over two nearby OCR lines.  A complete
    multi-token description is useful evidence in that narrow shape, but it
    remains below the decisive band and therefore cannot defeat a confirmed
    date/amount match or survive an equally-ranked rival.
    """
    if len(description_tokens) < 3 or description_hits < 3:
        return -1
    coverage = description_hits / len(description_tokens)
    if coverage < 0.8:
        return -1
    return min(_IDENTITY_ONLY_CEILING, 7 + round(coverage * 2))


def _description_reference_terms(description: str) -> tuple[str, ...]:
    return tuple(
        match.group(1)
        for match in re.finditer(
            r"check\s*#\s*([0-9]{3,})", str(description or ""), re.I
        )
    )


def _line_score(text: str, evidence: ExpenseEvidence) -> int:
    amount = _amount_decimal(evidence.amount)
    implicit_reference_terms = _description_reference_terms(evidence.description)
    reference_hit = any(
        re.search(rf"\b{re.escape(term)}\b", text, re.I)
        for term in (*evidence.reference_terms, *implicit_reference_terms)
        if term
    )
    exact_amount_hit = _amount_matches(text, amount)
    recovered_total_hit = (
        not exact_amount_hit
        and _labeled_total_amount_stem_matches(text, amount)
    )
    amount_hit = exact_amount_hit or recovered_total_hit
    date_hit = _date_matches(text, evidence.expense_date)
    # OCR on the rendered scanner page may preserve the human-readable
    # month/day but lose the year, or turn the date into prose such as
    # ``April 24, 2025``. The transaction row is still identifiable when
    # the amount and payee are present; retain that weaker date evidence for
    # ranking instead of requiring an exact calendar token.
    if not date_hit:
        date_hit = _statement_row_date_matches(text, evidence.expense_date)
        if not date_hit:
            lowered = str(text or '').lower()
            parsed_date = None
            try:
                parsed_date = datetime.strptime(
                    str(evidence.expense_date), '%Y-%m-%d'
                ).date()
            except ValueError:
                pass
            if parsed_date is not None:
                date_hit = (
                    f'{parsed_date.strftime("%B")} {parsed_date.day}' in lowered
                    or f'{parsed_date.strftime("%b")} {parsed_date.day}' in lowered
                )
    line_tokens = _tokens(text)
    description_tokens = _tokens(evidence.description)
    description_hits = len(line_tokens & description_tokens)
    vendor_hits = len(line_tokens & _tokens(evidence.vendor_key.replace("_", " ")))
    if not amount_hit and not reference_hit:
        # A scan EG has written on loses its amount column to OCR: the pen
        # stroke crosses the digits and "10.59" comes back as "1.59" or noise.
        # Rather than fail closed on the whole row, accept the date plus a
        # substantial share of the payee as identity, scored below the
        # decisive band so any rival row still cancels the box.
        # OCR routinely truncates the payee ("RADISSON HO"), so one hit on the
        # payee's own name says more than two on the city it shares with every
        # neighbouring row.
        head_hits = len(line_tokens & _payee_head_tokens(evidence.description))
        tail_hits = description_hits - head_hits
        if _complete_date_matches(text, evidence.expense_date):
            identified = description_hits >= 2
        elif _statement_row_date_matches(text, evidence.expense_date):
            # A bare M/D is weaker evidence than a printed calendar date, so
            # require the payee's head: geography alone must never win a row.
            identified = head_hits >= 1
        else:
            identified = False
        if identified:
            return min(
                _IDENTITY_ONLY_CEILING,
                5 + 2 * min(head_hits, 2) + min(tail_hits, 2)
                + min(vendor_hits, 2),
            )
        # If OCR has a transaction-like date at the start, it must be the
        # target date before a description-only fallback can be considered.
        # Otherwise a neighboring row with the same description but a
        # different date can tie the real row (the date may also occur later
        # in prose, such as ``2/15 sermon``).
        has_leading_date = _has_leading_calendar_date(text)
        description_only = (
            -1
            if has_leading_date
            else _description_only_score(description_tokens, description_hits)
        )
        strong_description = (
            -1
            if has_leading_date
            else _strong_description_score(description_tokens, description_hits)
        )
        return max(strong_description, description_only)
    score = 12 if reference_hit else 0
    if amount_hit:
        score += 5
    if recovered_total_hit:
        # Explicit TOTAL plus all but the final cent digit is stronger than an
        # exact authorization-copy amount with no expense semantics.
        score += 2
        if re.match(r"^\s*total\b", text, re.I):
            # Prefer the actual TOTAL row over a context window which merely
            # contains that row's text but draws a neighbouring region.
            score += 1
    if date_hit:
        score += 5
    score += min(description_hits, 3)
    score += min(vendor_hits, 2)
    # A candidate containing the target amount plus another decimal amount is
    # usually a generated window spanning two adjacent statement rows. Keep
    # the exact target row ahead of that composite even when the window has
    # slightly more description-token overlap.
    score -= 4 * _foreign_amount_count(text, amount)
    return score


def _amount_confirmed(text: str, evidence: ExpenseEvidence) -> bool:
    """Whether this line proves the amount itself, not just the expense's identity."""
    amount = _amount_decimal(evidence.amount)
    return _amount_matches(text, amount) or _labeled_total_amount_stem_matches(
        text, amount
    )


def _region_bounds(region: object) -> tuple[object, float, float] | None:
    """``(page, top, bottom)`` for a candidate region, or ``None`` if unknown."""
    if not isinstance(region, tuple) or not region:
        return None
    if len(region) == 4 and all(
        isinstance(value, (int, float)) for value in region
    ):
        # Image strategy: (left, top, right, bottom) on a single-page image.
        return "", float(region[1]), float(region[3])
    head, rest = region[0], region[1:]
    if len(rest) == 1 and all(hasattr(rest[0], side) for side in ("y0", "y1")):
        # PDF strategy: (page index, Rect).
        return head, float(rest[0].y0), float(rest[0].y1)
    if len(rest) == 3 and isinstance(rest[0], (int, float)):
        # Excel strategy: (sheet, row number, first column, last column).
        return repr(head), float(rest[0]), float(rest[0])
    return None


def _region_width(region: object) -> float | None:
    """Return the horizontal width of an image/PDF candidate, when available."""
    if (
        isinstance(region, tuple)
        and len(region) == 4
        and all(isinstance(value, (int, float)) for value in region)
    ):
        return float(region[2] - region[0])
    if (
        isinstance(region, tuple)
        and len(region) == 2
        and hasattr(region[1], "x0")
        and hasattr(region[1], "x1")
    ):
        return float(region[1].x1 - region[1].x0)
    return None


def _same_row(first: object, second: object) -> bool:
    """Whether two candidates are one physical row enumerated twice.

    Every strategy describes a row more than once on purpose — word-grouped
    and spatially reconstructed, plus an amount/date pair for PDFs — so two
    top-ranked candidates are usually the same row seen twice.  Only a tie
    between genuinely different rows is the ambiguity the tie rule exists for.
    """
    left = _region_bounds(first)
    right = _region_bounds(second)
    if left is None or right is None:
        return first == second
    if left[0] != right[0]:
        return False
    overlap = min(left[2], right[2]) - max(left[1], right[1])
    span = min(left[2] - left[1], right[2] - right[1])
    if span <= 0:
        return overlap >= 0
    return overlap > span / 2


def _regions_share_visual_row(first: object, second: object) -> bool:
    """Return whether two format-specific regions occupy one visual row.

    ``_same_row`` intentionally treats partially overlapping candidates as the
    same row for de-duplication. Row expansion also needs to join adjacent
    fragments whose edges merely touch, such as separate date, description,
    and amount cells.
    """
    left = _region_bounds(first)
    right = _region_bounds(second)
    if left is None or right is None or left[0] != right[0]:
        return False
    left_height = max(1.0, left[2] - left[1])
    right_height = max(1.0, right[2] - right[1])
    gap = max(left[1], right[1]) - min(left[2], right[2])
    return gap <= max(3.0, min(left_height, right_height) * 0.35)


def _regions_are_adjacent_visual_row(first: object, second: object) -> bool:
    """Return whether nearby fragments may be one wrapped/card transaction.

    A statement-card scan can place the date on a line above the merchant and
    amount. This uses a height-relative gap so PDF and image strategies share
    the same rule; callers still require independent row evidence before
    joining anything.
    """
    left = _region_bounds(first)
    right = _region_bounds(second)
    if left is None or right is None or left[0] != right[0]:
        return False
    left_height = max(1.0, left[2] - left[1])
    right_height = max(1.0, right[2] - right[1])
    gap = max(left[1], right[1]) - min(left[2], right[2])
    return gap <= max(3.0, min(left_height, right_height) * 2.25)


def _region_union(first: object, second: object) -> object | None:
    """Union two image/PDF regions without knowing their concrete adapter."""
    if (
        isinstance(first, tuple)
        and isinstance(second, tuple)
        and len(first) == 4
        and len(second) == 4
        and all(isinstance(value, (int, float)) for value in first + second)
    ):
        return (
            min(first[0], second[0]),
            min(first[1], second[1]),
            max(first[2], second[2]),
            max(first[3], second[3]),
        )
    if (
        isinstance(first, tuple)
        and isinstance(second, tuple)
        and len(first) == 2
        and len(second) == 2
        and first[0] == second[0]
        and all(hasattr(first[1], side) for side in ("x0", "y0", "x1", "y1"))
        and all(hasattr(second[1], side) for side in ("x0", "y0", "x1", "y1"))
    ):
        rect_type = type(first[1])
        return (
            first[0],
            rect_type(
                min(first[1].x0, second[1].x0),
                min(first[1].y0, second[1].y0),
                max(first[1].x1, second[1].x1),
                max(first[1].y1, second[1].y1),
            ),
        )
    return None


def _row_fragment_supports_expansion(
    text: str,
    evidence: ExpenseEvidence,
) -> bool:
    """Whether a fragment independently identifies this expense row."""
    line_tokens = _tokens(text)
    description_tokens = _tokens(evidence.description)
    description_hits = line_tokens & description_tokens
    head_hits = line_tokens & _payee_head_tokens(evidence.description)
    vendor_hits = line_tokens & _tokens(evidence.vendor_key.replace("_", " "))
    # A merchant head or vendor key is distinctive enough on its own. For a
    # description without a merchant head, require two description tokens so a
    # shared city/state token cannot widen a box into a neighboring row.
    return bool(head_hits or vendor_hits or len(description_hits) >= 2)


def _expand_selected_row_region(
    lines: Sequence[tuple[str, object]],
    selected_region: object,
    evidence: ExpenseEvidence,
    selected_text: str = "",
) -> tuple[object, str]:
    """Widen a winning candidate to matching same-row evidence fragments.

    OCR/PDF adapters can enumerate one printed row as separate date,
    description, and amount fragments. Selecting the right date or amount can
    therefore still produce a red box around only one cell. Expand only to
    fragments that independently match the known description/vendor or exact
    amount; never expand a neighboring transaction merely because it is close.
    """
    expanded = selected_region
    supporting_text: list[str] = []
    selected_bounds = _region_bounds(selected_region)
    if selected_bounds is None:
        return selected_region, ""
    selected_height = max(1.0, selected_bounds[2] - selected_bounds[1])
    selected_width = max(1.0, _region_width(selected_region) or 0.0)
    max_fragment_height = max(
        selected_height * 1.5,
        selected_height + min(40.0, selected_height * 0.5),
    )
    # A broad candidate that already contains the merchant and exact amount is
    # complete. Do not let nearby summary/card text widen it into the next
    # transaction; this is common when OCR emits overlapping page-level lines.
    selected_text = selected_text or next(
        (str(text) for text, region in lines if region == selected_region),
        "",
    )
    description_tokens = _tokens(evidence.description)
    description_hits = _tokens(selected_text) & description_tokens
    description_is_complete = bool(
        description_tokens
        and len(description_hits) >= max(
            3,
            round(len(description_tokens) * 0.75),
        )
    )
    # A reconstructed OCR/window candidate can contain the full row's text
    # while its geometry still covers only the date cell. Treat only a
    # genuinely broad region as complete; narrow amount-confirmed winners must
    # continue through same-row fragment expansion below.
    selected_region_is_broad = selected_width >= max(800.0, selected_height * 12.0)
    selected_text_has_amount = _amount_confirmed(selected_text, evidence)
    selected_text_has_description = _row_fragment_supports_expansion(
        selected_text, evidence
    )
    image_region = (
        isinstance(selected_region, tuple)
        and len(selected_region) == 4
        and all(isinstance(value, (int, float)) for value in selected_region)
    )
    date_fragment_present = image_region and any(
        _regions_are_adjacent_visual_row(selected_region, region)
        and _date_matches(str(text), evidence.expense_date)
        for text, region in lines
    )
    if (
        selected_region_is_broad
        and _row_fragment_supports_expansion(selected_text, evidence)
        and (
            _amount_confirmed(selected_text, evidence)
            or (
                description_is_complete
            )
        )
    ):
        # A broad OCR candidate that already contains the complete merchant
        # description is the row/card itself. Do not union it with nearby
        # repeated Diners Club/card fragments; those are separate statement
        # sections, not missing cells in this transaction.
        return selected_region, ""
    amount_pattern = re.compile(r"(?<!\d)\$?\d[\d,]*\.\d{2}(?!\d)")
    description_fragment_present = any(
        _regions_are_adjacent_visual_row(selected_region, region)
        and (
            _regions_share_visual_row(selected_region, region)
            or (
                (bounds := _region_bounds(region)) is not None
                and bounds[2] - bounds[1] <= max_fragment_height
            )
        )
        and _row_fragment_supports_expansion(text, evidence)
        for text, region in lines
    )
    for text, region in lines:
        if not _regions_are_adjacent_visual_row(selected_region, region):
            continue
        # Fragments must share the same visual row, not merely be "adjacent".
        # A complete neighboring transaction row must never be pulled in just
        # because it contains the same merchant name. Only same-row fragments
        # (date/description/amount cells of the selected transaction) or
        # wrapped multi-line card descriptions should be joined.
        if not _regions_share_visual_row(selected_region, region):
            # Allow date fragments above the selected region when the date is
            # on a separate line (card statements). Otherwise, only join
            # fragments that truly share the same visual row.
            region_bounds = _region_bounds(region)
            if region_bounds is None:
                continue
            is_date_above = (
                date_fragment_present
                and region_bounds[1] < selected_bounds[1]
                and _date_matches(str(text), evidence.expense_date)
            )
            if not is_date_above:
                continue
        region_top = _region_bounds(region)[1] if _region_bounds(region) else None
        if date_fragment_present and region_top is not None and region_top < selected_bounds[1]:
            # Already handled above - skip date fragments above unless they're the date
            if not _date_matches(str(text), evidence.expense_date):
                continue
        supports_description = _row_fragment_supports_expansion(text, evidence)
        supports_date = (
            date_fragment_present
            and _date_matches(str(text), evidence.expense_date)
        )
        amount_hits = amount_pattern.findall(str(text))
        region_bounds = _region_bounds(region)
        if region_bounds is None:
            continue
        region_height = max(1.0, region_bounds[2] - region_bounds[1])
        if region_height > max_fragment_height:
            continue
        supports_amount = (
            description_fragment_present
            and _amount_confirmed(str(text), evidence)
            and len(amount_hits) <= 1
            and region_height <= selected_height * 3
        )
        if not supports_description and not supports_amount and not supports_date:
            continue
        union = _region_union(expanded, region)
        if union is not None:
            expanded = union
            supporting_text.append(str(text))
    return expanded, " ".join(supporting_text)


def _evidence_description_windows(
    lines: Sequence[tuple[str, object]],
    evidence: ExpenseEvidence,
) -> list[tuple[str, object]]:
    """Join adjacent OCR fragments when they complete this description.

    Some statement-summary cards put the charge label on one visual line and
    the card/vendor identifier on the next.  The Diners Club scan does this for
    ``A)INUAL FEE`` and ``Diners Club | x-0587``.  The generic wrapped-line
    window intentionally rejects that gap, so add an evidence-aware candidate
    only when both fragments contribute description tokens and their union
    covers most of the expected description.  This avoids joining ordinary
    neighboring transaction rows merely because they are close together.
    """
    description_tokens = _tokens(evidence.description)
    if len(description_tokens) < 2:
        return []
    candidates: list[tuple[str, object]] = []
    for index, (first_text, first_region) in enumerate(lines):
        first_bounds = _region_bounds(first_region)
        if first_bounds is None or not (
            isinstance(first_region, tuple)
            and len(first_region) == 4
            and all(isinstance(value, (int, float)) for value in first_region)
        ):
            continue
        for second_text, second_region in lines[index + 1:]:
            second_bounds = _region_bounds(second_region)
            if second_bounds is None or first_bounds[0] != second_bounds[0]:
                continue
            if second_bounds[1] < first_bounds[1]:
                continue
            gap = second_bounds[1] - first_bounds[2]
            first_height = first_bounds[2] - first_bounds[1]
            second_height = second_bounds[2] - second_bounds[1]
            if gap > max(12.0, min(first_height, second_height) * 1.5):
                continue
            horizontal_overlap = min(first_region[2], second_region[2]) - max(
                first_region[0], second_region[0]
            )
            if horizontal_overlap <= 0:
                continue
            first_hits = _tokens(first_text) & description_tokens
            second_hits = _tokens(second_text) & description_tokens
            combined_hits = first_hits | second_hits
            if not first_hits or not second_hits:
                continue
            if len(combined_hits) / len(description_tokens) < 0.6:
                continue
            region = (
                min(first_region[0], second_region[0]),
                min(first_region[1], second_region[1]),
                max(first_region[2], second_region[2]),
                max(first_region[3], second_region[3]),
            )
            candidates.append((f"{first_text} {second_text}", region))
    return candidates


def _best_line(
    lines: Iterable[tuple[str, object]],
    evidence: ExpenseEvidence,
) -> tuple[object | None, int, str]:
    """Return ``(region, score, text)`` for the winning line, or ``(None, …)``."""
    candidate_lines = list(lines)
    # Composite windows are only a recovery path for wrapped rows. They can
    # carry a correct amount/date copied from a nearby row, so never let one
    # outrank an original OCR line with the same score.
    original_line_count = len(candidate_lines)
    candidate_lines.extend(_evidence_description_windows(candidate_lines, evidence))
    def _rank_key(item: tuple[int, object, str, bool]) -> tuple[object, ...]:
        score, region, text, is_original = item
        # When OCR lost both date and amount, several overlapping page/card
        # candidates may have identical identity scores. Prefer the most
        # precise candidate in that case: a narrow description/card row is
        # evidence, while a larger composite also includes unrelated header
        # or category text. Confirmed amount/date candidates retain priority.
        amount_confirmed = _amount_confirmed(text, evidence)
        date_confirmed = _date_matches(text, evidence.expense_date)
        description_hits = len(_tokens(text) & _tokens(evidence.description))
        identity_only_precision = not amount_confirmed and not date_confirmed
        # Objective signals break ties, but the original score remains the
        # primary rank: a description-only merchant row can legitimately be
        # stronger than a generic statement line carrying the same amount.
        # Within equal scores, amount/date evidence wins over broad windows.
        evidence_strength = (
            2 if amount_confirmed and date_confirmed
            else 1 if amount_confirmed or date_confirmed
            else 0
        )
        return (
            score,
            evidence_strength,
            amount_confirmed,
            date_confirmed,
            description_hits,
            -_region_area(region) if identity_only_precision else 0.0,
            is_original,
        )

    ranked = sorted(
        (
            (_line_score(text, evidence), region, text, index < original_line_count)
            for index, (text, region) in enumerate(candidate_lines)
        ),
        key=_rank_key,
        reverse=True,
    )
    if not ranked:
        return None, -1, ""
    if ranked[0][0] < 7:
        amount = _amount_decimal(evidence.amount)
        matches = [item for item in ranked if _amount_matches(item[2], amount)]
        if matches:
            parents = list(range(len(matches)))

            def _root(i: int) -> int:
                while parents[i] != i:
                    parents[i] = parents[parents[i]]
                    i = parents[i]
                return i

            for i in range(len(matches)):
                for j in range(i + 1, len(matches)):
                    if _same_row(matches[i][1], matches[j][1]):
                        root_i, root_j = _root(i), _root(j)
                        if root_i != root_j:
                            parents[root_i] = root_j
            if len({_root(i) for i in range(len(matches))}) == 1:
                return matches[0][1], matches[0][0], matches[0][2]
            if all(_TOTAL_WORD_RE.search(item[2]) for item in matches):
                best = max(
                    matches,
                    key=lambda item: (_region_bounds(item[1]) or ("", 0.0, 0.0))[2],
                )
                return best[1], best[0], best[2]
        return None, ranked[0][0], ""
    # A date+amount match is decisive.  For looser description matches, reject
    # a tie instead of boxing the wrong repeated amount.
    rival = next(
        (
            item
            for item in ranked[1:]
            if not _same_row(item[1], ranked[0][1])
            and not (item[3] is False and ranked[0][3] is True)
        ),
        None,
    )
    if (
        ranked[0][0] < _DECISIVE_SCORE
        and rival is not None
        and rival[0] == ranked[0][0]
    ):
        return None, ranked[0][0], ""
    # Even when score >= _DECISIVE_SCORE, reject ties between rows that have
    # the same date+amount but only weak description overlap. A score of 10+
    # from date+amount+vendor with minimal distinctive description tokens is
    # still ambiguous when multiple rows share that pattern.
    if rival is not None and rival[0] == ranked[0][0]:
        winner_description_hits = len(
            _tokens(ranked[0][2]) & _tokens(evidence.description)
        )
        rival_description_hits = len(
            _tokens(rival[2]) & _tokens(evidence.description)
        )
        # Only accept the winner if it has meaningfully more description
        # overlap than its rival (at least 2 more tokens). Otherwise, the
        # common tokens (like "store" in "First Store" vs "Second Store")
        # don't disambiguate them.
        if winner_description_hits < rival_description_hits + 2:
            # OCR candidates generated from a single visual card often have
            # different vertical extents but the same transaction text.
            # Treat overlapping candidates as one row (the normal path), and
            # prefer the tightest region so the fail-closed tie rule does not
            # discard a valid card row merely because a window included its
            # neighboring labels.
            if _same_row(ranked[0][1], rival[1]):
                tightest = min(
                    (ranked[0], rival),
                    key=lambda item: _region_area(item[1]),
                )
                return tightest[1], tightest[0], tightest[2]
            # Before failing closed, check if the winner and rival are
            # essentially the same transaction with just repeated amounts.
            # This happens when OCR emits the same row multiple times or
            # when the amount appears in multiple columns.
            amount = _amount_decimal(evidence.amount)
            if amount is not None:
                winner_amount_count = len(re.findall(
                    rf"(?<![\d.])\$?\s*{re.escape(f'{amount:.2f}')}(?!\d)",
                    ranked[0][2]
                ))
                rival_amount_count = len(re.findall(
                    rf"(?<![\d.])\$?\s*{re.escape(f'{amount:.2f}')}(?!\d)",
                    rival[2]
                ))
                # If both have the same description hits and only differ in
                # how many times the amount appears, prefer the one with
                # fewer repetitions (likely more precise).
                if winner_description_hits == rival_description_hits:
                    if winner_amount_count > 0 and winner_amount_count < rival_amount_count:
                        return ranked[0][1], ranked[0][0], ranked[0][2]
                    if rival_amount_count > 0 and rival_amount_count < winner_amount_count:
                        return rival[1], rival[0], rival[2]
            # True ambiguity: fail closed
            return None, ranked[0][0], ""
    return ranked[0][1], ranked[0][0], ranked[0][2]


def _spatial_lines(
    words: Iterable[tuple[str, float, float, float, float]],
) -> list[tuple[str, tuple[float, float, float, float]]]:
    """Reconstruct visual rows when document extractors split table columns."""
    clean = [
        (str(text).strip(), float(x0), float(y0), float(x1), float(y1))
        for text, x0, y0, x1, y1 in words
        if str(text).strip()
    ]
    if not clean:
        return []
    heights = [max(1.0, word[4] - word[2]) for word in clean]
    tolerance = max(3.0, statistics.median(heights) * 0.8)
    rows: list[list[tuple[str, float, float, float, float]]] = []
    centers: list[float] = []
    for word in sorted(clean, key=lambda item: ((item[2] + item[4]) / 2, item[1])):
        center = (word[2] + word[4]) / 2
        best = None
        best_distance = None
        for index, row_center in enumerate(centers):
            distance = abs(center - row_center)
            if distance <= tolerance and (
                best_distance is None or distance < best_distance
            ):
                best = index
                best_distance = distance
        if best is None:
            rows.append([word])
            centers.append(center)
        else:
            rows[best].append(word)
            centers[best] = sum(
                (item[2] + item[4]) / 2 for item in rows[best]
            ) / len(rows[best])
    output = []
    for row in rows:
        row.sort(key=lambda item: item[1])
        output.append(
            (
                " ".join(item[0] for item in row),
                (
                    min(item[1] for item in row),
                    min(item[2] for item in row),
                    max(item[3] for item in row),
                    max(item[4] for item in row),
                ),
            )
        )
    return output


def _wrapped_line_windows(
    lines: Sequence[tuple[str, tuple[float, float, float, float]]],
) -> list[tuple[str, tuple[float, float, float, float]]]:
    """Join overlapping OCR lines from one wrapped table row.

    Tesseract can split a wrapped description into two rows and attach the
    date/amount noise to only one of them. Only vertically touching lines with
    overlapping horizontal bounds are joined, so adjacent table rows remain
    separate candidates and the normal ambiguity rules still apply.
    """
    clean = [
        (
            str(text).strip(),
            tuple(float(value) for value in region),
        )
        for text, region in lines
        if str(text).strip()
        and len(region) == 4
        and all(isinstance(value, (int, float)) for value in region)
    ]
    if len(clean) < 2:
        return []
    heights = [
        max(1.0, region[3] - region[1])
        for _text, region in clean
    ]
    max_gap = max(2.0, statistics.median(heights) * 0.25)
    ordered = sorted(clean, key=lambda item: (item[1][1], item[1][0]))
    windows = []
    for index, (text, region) in enumerate(ordered[:-1]):
        parts = [text]
        left, top, right, bottom = region
        for next_text, next_region in ordered[index + 1:index + 3]:
            horizontal_overlap = min(right, next_region[2]) - max(
                left, next_region[0]
            )
            if next_region[1] > bottom + max_gap or horizontal_overlap <= 0:
                break
            parts.append(next_text)
            left = min(left, next_region[0])
            top = min(top, next_region[1])
            right = max(right, next_region[2])
            bottom = max(bottom, next_region[3])
            windows.append((" ".join(parts), (left, top, right, bottom)))
    return windows


def _bill_summary_windows(
    lines: Sequence[tuple[str, tuple[float, float, float, float]]],
) -> list[tuple[str, tuple[float, float, float, float]]]:
    """Join three close rows when a bill prints charge and date separately.

    Matching remains fail-closed: this only supplies candidates. The normal
    scorer still requires the target amount and date before selecting one.
    """
    clean = sorted(
        (
            (str(text).strip(), tuple(float(value) for value in region))
            for text, region in lines
            if str(text).strip() and len(region) == 4
        ),
        key=lambda item: (item[1][1], item[1][0]),
    )
    output = []
    for index in range(len(clean) - 2):
        window = clean[index:index + 3]
        heights = [max(1.0, region[3] - region[1]) for _text, region in window]
        max_gap = max(12.0, statistics.median(heights) * 0.75)
        if any(
            window[offset + 1][1][1] - window[offset][1][3] > max_gap
            for offset in range(2)
        ):
            continue
        # Find which row contains the expense amount — that is the line to box.
        # For DTE bills (expense 1985), the amount is in the first row and the
        # date is in a later row. For Priority Health EOBs (expense 1988), the
        # date and description are in the first row but the amount is in a later
        # row. Box whichever row carries the expense amount, not the context.
        # Prefer rows with charge/billed/payment labels over balance summaries.
        expense_labels = re.compile(r"\b(?:billed|charge|payment|paid|due)\b", re.I)
        balance_labels = re.compile(r"\b(?:balance|total\s+balance|account\s+balance)\b", re.I)

        # First, try rows with explicit expense labels and amounts
        amount_row_index = next(
            (
                i
                for i, (text, _region) in enumerate(window)
                if re.search(r"(?:^|\s)\$\s*\d+[,\d]*\.\d{2}(?:\s|$)", text)
                and expense_labels.search(text)
                and not balance_labels.search(text)
            ),
            None,
        )
        if amount_row_index is None:
            # Second, try any row with amount that's NOT a balance
            amount_row_index = next(
                (
                    i
                    for i, (text, _region) in enumerate(window)
                    if re.search(r"(?:^|\s)[\$]?\d+[,\d]*\.\d{2}(?:\s|$)", text)
                    and not balance_labels.search(text)
                    and not re.search(r"\d{4}", text)  # Exclude dates like "2025"
                ),
                0,
            )
        expense_region = window[amount_row_index][1]
        output.append(
            (" ".join(text for text, _region in window), expense_region)
        )
    return output


def _image_expense_candidates(
    ocr_data: dict[str, Sequence[object]],
) -> list[tuple[str, tuple[float, float, float, float]]]:
    """Pure OCR-to-candidate transform used by the image adapter."""
    grouped: dict[tuple[int, int, int], list[int]] = {}
    spatial_words = []
    for index, text in enumerate(ocr_data.get("text", [])):
        if not str(text or "").strip():
            continue
        key = (
            int(ocr_data["block_num"][index]),
            int(ocr_data["par_num"][index]),
            int(ocr_data["line_num"][index]),
        )
        grouped.setdefault(key, []).append(index)
        left = int(ocr_data["left"][index])
        top = int(ocr_data["top"][index])
        spatial_words.append(
            (
                str(text),
                left,
                top,
                left + int(ocr_data["width"][index]),
                top + int(ocr_data["height"][index]),
            )
        )

    grouped_lines = []
    for indexes in grouped.values():
        text = " ".join(str(ocr_data["text"][index]) for index in indexes)
        grouped_lines.append(
            (
                text,
                (
                    min(int(ocr_data["left"][index]) for index in indexes),
                    min(int(ocr_data["top"][index]) for index in indexes),
                    max(
                        int(ocr_data["left"][index])
                        + int(ocr_data["width"][index])
                        for index in indexes
                    ),
                    max(
                        int(ocr_data["top"][index])
                        + int(ocr_data["height"][index])
                        for index in indexes
                    ),
                ),
            )
        )
    spatial_lines = _spatial_lines(spatial_words)
    candidates = grouped_lines + _bill_summary_windows(grouped_lines)
    candidates.extend(_wrapped_line_windows(grouped_lines))
    candidates.extend(spatial_lines)
    candidates.extend(_bill_summary_windows(spatial_lines))
    candidates.extend(_wrapped_line_windows(spatial_lines))
    return candidates


class PdfExpenseDocumentAnnotator(IExpenseDocumentAnnotator):
    def supports(self, source_path: str) -> bool:
        return Path(source_path).suffix.lower() == ".pdf"

    def annotate(
        self,
        source_path: str,
        output_path: str,
        evidence: ExpenseEvidence,
    ) -> AnnotationResult:
        try:
            import fitz
        except ImportError:
            return AnnotationResult(source_path, False, "PyMuPDF is not installed.")

        document = fitz.open(source_path)
        candidates: list[tuple[str, tuple[int, object]]] = []
        try:
            for page_index, page in enumerate(document):
                page_words = list(page.get_text("words"))
                amount = _amount_decimal(evidence.amount)
                median_height = statistics.median(
                    max(1.0, word[3] - word[1]) for word in page_words
                ) if page_words else 1.0
                row_tolerance = max(3.0, median_height * 0.8)
                # Table PDFs often encode each cell as a separate "line".
                # Pair the exact amount with the nearest same-row date first;
                # its tight rectangle avoids boxing three side-by-side checks
                # that happen to share one visual y coordinate.
                for amount_word in page_words:
                    if not _amount_matches(str(amount_word[4]), amount):
                        continue
                    amount_center_y = (amount_word[1] + amount_word[3]) / 2
                    same_row = [
                        word
                        for word in page_words
                        if abs(
                            ((word[1] + word[3]) / 2) - amount_center_y
                        ) <= row_tolerance
                    ]
                    dates = [
                        word for word in same_row
                        if _date_matches(str(word[4]), evidence.expense_date)
                    ]
                    if not dates:
                        continue
                    date_word = min(
                        dates,
                        key=lambda word: abs(word[0] - amount_word[0]),
                    )
                    left = min(date_word[0], amount_word[0])
                    right = max(date_word[2], amount_word[2])
                    nearby = [
                        word for word in same_row
                        if left - 75 <= word[0] and word[2] <= right + 8
                    ]
                    rect = fitz.Rect(
                        min(word[0] for word in nearby),
                        min(word[1] for word in nearby),
                        max(word[2] for word in nearby),
                        max(word[3] for word in nearby),
                    )
                    text = " ".join(
                        str(word[4]) for word in sorted(
                            nearby, key=lambda word: word[0])
                    )
                    candidates.append((text, (page_index, rect)))
                grouped: dict[tuple[int, int], list[tuple]] = {}
                for word in page_words:
                    grouped.setdefault((int(word[5]), int(word[6])), []).append(word)
                for words in grouped.values():
                    words.sort(key=lambda word: word[0])
                    text = " ".join(str(word[4]) for word in words)
                    rect = fitz.Rect(
                        min(word[0] for word in words),
                        min(word[1] for word in words),
                        max(word[2] for word in words),
                        max(word[3] for word in words),
                    )
                    candidates.append((text, (page_index, rect)))
                for text, bounds in _spatial_lines(
                    (word[4], word[0], word[1], word[2], word[3])
                    for word in page_words
                ):
                    candidates.append(
                        (text, (page_index, fitz.Rect(*bounds)))
                    )
            region, _score, _text = _best_line(candidates, evidence)
            if region is None:
                return AnnotationResult(
                    source_path,
                    False,
                    "No high-confidence expense row was found in the PDF.",
                )
            page_index, rect = region
            expanded_region, expanded_text = _expand_selected_row_region(
                candidates,
                region,
                evidence,
                _text,
            )
            if expanded_text:
                page_index, rect = expanded_region
            page = document[page_index]
            rect = fitz.Rect(
                max(page.rect.x0, rect.x0 - 4),
                max(page.rect.y0, rect.y0 - 3),
                min(page.rect.x1, rect.x1 + 4),
                min(page.rect.y1, rect.y1 + 3),
            )
            page.draw_rect(rect, color=(1, 0, 0), width=3, overlay=True)
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            document.save(output_path, garbage=3, deflate=True)
            return AnnotationResult(output_path, True, page=page_index + 1)
        finally:
            document.close()


class PdfCheckNumberResolver(IExpenseReferenceResolver):
    """Find the check number beside this expense in a related bank statement."""

    def resolve(self, evidence: ExpenseEvidence) -> tuple[str, ...]:
        source_path = evidence.related_document_path
        if not source_path or Path(source_path).suffix.lower() != ".pdf":
            return ()
        try:
            import fitz
        except ImportError:
            return ()
        amount = _amount_decimal(evidence.amount)
        try:
            document = fitz.open(source_path)
        except Exception:
            return ()
        try:
            for page in document:
                words = list(page.get_text("words"))
                if not words:
                    continue
                tolerance = max(
                    3.0,
                    statistics.median(
                        max(1.0, word[3] - word[1]) for word in words
                    ) * 0.8,
                )
                for amount_word in words:
                    if not _amount_matches(str(amount_word[4]), amount):
                        continue
                    center_y = (amount_word[1] + amount_word[3]) / 2
                    row = [
                        word for word in words
                        if abs(((word[1] + word[3]) / 2) - center_y)
                        <= tolerance
                    ]
                    dates = [
                        word for word in row
                        if _date_matches(str(word[4]), evidence.expense_date)
                    ]
                    if not dates:
                        continue
                    date_word = min(
                        dates,
                        key=lambda word: abs(word[0] - amount_word[0]),
                    )
                    identifiers = [
                        str(word[4])
                        for word in row
                        if word[2] <= date_word[0]
                        and date_word[0] - word[2] <= 95
                        and re.fullmatch(r"\d{4,6}", str(word[4]))
                    ]
                    if identifiers:
                        return (identifiers[-1],)
        finally:
            document.close()
        return ()


class ImageExpenseDocumentAnnotator(IExpenseDocumentAnnotator):
    _EXTENSIONS = {".jpg", ".jpeg", ".png", ".bmp", ".tif", ".tiff", ".webp"}
    _MIN_FALLBACK_CONFIDENCE = 0.9

    def __init__(
        self,
        reference_resolver: IExpenseReferenceResolver | None = None,
        fallback_matcher: IImageRegionFallbackMatcher | None = None,
    ) -> None:
        self._reference_resolver = reference_resolver
        self._fallback_matcher = fallback_matcher

    def supports(self, source_path: str) -> bool:
        return Path(source_path).suffix.lower() in self._EXTENSIONS

    def annotate(
        self,
        source_path: str,
        output_path: str,
        evidence: ExpenseEvidence,
    ) -> AnnotationResult:
        try:
            from PIL import Image, ImageDraw
            import pytesseract
            from pytesseract import Output
        except ImportError:
            return AnnotationResult(
                source_path, False, "Pillow or pytesseract is not installed."
            )

        try:
            image = Image.open(source_path)
            # Give Tesseract the original file so scan DPI/encoding metadata is
            # preserved. Passing a decoded PIL object makes pytesseract write a
            # temporary image; on the DTE bill that erased both 53.06 readings.
            # PSM 1 enables Tesseract's orientation detection before layout
            # analysis. Scanner hardware may deliver an otherwise valid form
            # rotated 180 degrees (Children's Vision, expense 1987).
            ocr_candidates = []
            for config in ("--psm 1", "--psm 6"):
                data = pytesseract.image_to_data(
                    source_path,
                    output_type=Output.DICT,
                    config=config,
                )
                ocr_candidates.extend(_image_expense_candidates(data))
        except Exception as exc:
            return AnnotationResult(
                source_path, False, f"Image OCR was unavailable: {exc}"
            )

        match_evidence = evidence
        if self._reference_resolver is not None:
            terms = self._reference_resolver.resolve(evidence)
            if terms:
                match_evidence = replace(evidence, reference_terms=terms)

        lines = ocr_candidates
        region, _score, matched_text = _best_line(lines, match_evidence)
        used_fallback = False
        if region is None and self._fallback_matcher is not None:
            try:
                fallback_match = self._fallback_matcher.find_region(
                    source_path,
                    match_evidence,
                )
            except Exception as exc:
                return AnnotationResult(
                    source_path,
                    False,
                    f"Image fallback matching was unavailable: {exc}",
                )
            if (
                fallback_match is not None
                and isinstance(fallback_match.confidence, (int, float))
                and not isinstance(fallback_match.confidence, bool)
                and math.isfinite(fallback_match.confidence)
                and fallback_match.confidence >= self._MIN_FALLBACK_CONFIDENCE
                and fallback_match.confidence <= 1
                and _valid_image_region(
                    fallback_match.region,
                    image.width,
                    image.height,
                )
            ):
                region = fallback_match.region
                used_fallback = True
        if region is None:
            return AnnotationResult(
                source_path,
                False,
                "No high-confidence expense row was found in the image.",
            )
        if not used_fallback:
            expanded_region, expanded_text = _expand_selected_row_region(
                lines,
                region,
                match_evidence,
                matched_text,
            )
            if expanded_text:
                region = expanded_region
        left, top, right, bottom = region
        if (
            not used_fallback
            and match_evidence.reference_terms
            and not _amount_confirmed(matched_text, match_evidence)
        ):
            # A ledger's related statement may identify the expense by check
            # number even when cross-outs defeat OCR on the payment line.
            # Include the payment/confirmation lines immediately above it.
            left = min(left, round(image.width * 0.08))
            right = max(right, round(image.width * 0.94))
            top = max(0, top - round(image.height * 0.06))
            bottom = min(image.height - 1, bottom + round(image.height * 0.01))
        elif (
            not used_fallback
            and not _amount_confirmed(matched_text, match_evidence)
        ):
            # Identity-only match: OCR lost the amount column, so the winning
            # line stops short of it.  Box the full statement row so the red
            # box still shows EG where the money came from.
            description_tokens = _tokens(match_evidence.description)
            description_hits = len(
                _tokens(matched_text) & description_tokens
            )
            if _strong_description_score(
                description_tokens, description_hits
            ) < 0:
                left = min(left, round(image.width * 0.08))
            right = max(right, round(image.width * 0.94))
        padding = max(8, round(image.width * 0.004))
        draw = ImageDraw.Draw(image)
        draw.rectangle(
            (
                max(0, left - padding),
                max(0, top - padding),
                min(image.width - 1, right + padding),
                min(image.height - 1, bottom + padding),
            ),
            outline="red",
            width=max(5, round(image.width * 0.003)),
        )
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        suffix = Path(output_path).suffix.lower()
        output_format = {
            ".jpg": "JPEG",
            ".jpeg": "JPEG",
            ".png": "PNG",
            ".bmp": "BMP",
            ".tif": "TIFF",
            ".tiff": "TIFF",
            ".webp": "WEBP",
        }.get(suffix, image.format or "PNG")
        if output_format == "JPEG" and image.mode not in ("RGB", "L"):
            image = image.convert("RGB")
        image.save(output_path, format=output_format)
        return AnnotationResult(output_path, True)


class ExcelExpenseDocumentAnnotator(IExpenseDocumentAnnotator):
    _EXTENSIONS = {".xlsx", ".xlsm"}

    def supports(self, source_path: str) -> bool:
        return Path(source_path).suffix.lower() in self._EXTENSIONS

    def annotate(
        self,
        source_path: str,
        output_path: str,
        evidence: ExpenseEvidence,
    ) -> AnnotationResult:
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Border, Side
        except ImportError:
            return AnnotationResult(source_path, False, "openpyxl is not installed.")

        keep_vba = Path(source_path).suffix.lower() == ".xlsm"
        workbook = load_workbook(source_path, keep_vba=keep_vba, data_only=False)
        candidates: list[tuple[str, tuple[object, int, int, int]]] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                populated = [cell for cell in row if cell.value not in (None, "")]
                if not populated:
                    continue
                text_parts = []
                for cell in populated:
                    value = cell.value
                    if isinstance(value, (datetime, date)):
                        text_parts.append(value.strftime("%m/%d/%Y"))
                    elif isinstance(value, (int, float, Decimal)):
                        text_parts.append(f"{value:.2f}")
                    else:
                        text_parts.append(str(value))
                candidates.append(
                    (
                        " ".join(text_parts),
                        (sheet, row[0].row, populated[0].column, populated[-1].column),
                    )
                )
        region, _score, _text = _best_line(candidates, evidence)
        if region is None:
            workbook.close()
            return AnnotationResult(
                source_path,
                False,
                "No high-confidence expense row was found in the spreadsheet.",
            )
        sheet, row_number, first_column, last_column = region
        red = Side(style="thick", color="FFFF0000")
        for column in range(first_column, last_column + 1):
            cell = sheet.cell(row=row_number, column=column)
            cell.border = Border(
                left=red if column == first_column else cell.border.left,
                right=red if column == last_column else cell.border.right,
                top=red,
                bottom=red,
            )
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        workbook.save(output_path)
        workbook.close()
        return AnnotationResult(output_path, True)


class ExpenseDocumentAnnotationService(IExpenseDocumentAnnotationService):
    def __init__(
        self,
        annotators: Sequence[IExpenseDocumentAnnotator],
        cache_dir: str,
    ) -> None:
        self._annotators = tuple(annotators)
        self._cache_dir = os.path.abspath(cache_dir)
        self._lock = threading.Lock()

    def prepare(
        self,
        source_path: str,
        evidence: ExpenseEvidence,
    ) -> AnnotationResult:
        source_path = os.path.abspath(source_path)
        annotator = next(
            (item for item in self._annotators if item.supports(source_path)), None
        )
        if annotator is None:
            return AnnotationResult(
                source_path, False, "This document format has no annotation strategy."
            )
        try:
            stat = os.stat(source_path)
        except OSError as exc:
            return AnnotationResult(source_path, False, f"Document unavailable: {exc}")
        fingerprint = hashlib.sha256(
            repr(
                (
                    source_path,
                    stat.st_mtime_ns,
                    stat.st_size,
                    evidence,
                    type(annotator).__name__,
                    ANNOTATION_SCHEMA_VERSION,
                )
            ).encode("utf-8")
        ).hexdigest()[:24]
        suffix = Path(source_path).suffix.lower()
        output_path = os.path.join(
            self._cache_dir,
            f"expense-{evidence.expense_id}-{fingerprint}{suffix}",
        )
        if os.path.isfile(output_path):
            return AnnotationResult(output_path, True)
        with self._lock:
            if os.path.isfile(output_path):
                return AnnotationResult(output_path, True)
            try:
                return annotator.annotate(source_path, output_path, evidence)
            except Exception as exc:
                return AnnotationResult(
                    source_path,
                    False,
                    f"Highlighting failed safely: {type(exc).__name__}: {exc}",
                )


def render_excel_for_browser(source_path: str, output_path: str) -> str:
    """Render an annotated workbook as a self-contained inline HTML view."""
    from openpyxl import load_workbook

    workbook = load_workbook(source_path, data_only=False)
    parts = [
        "<!doctype html><html><head><meta charset='utf-8'>",
        "<title>Supporting document</title>",
        "<style>body{font-family:Arial,sans-serif;margin:20px}" 
        "table{border-collapse:collapse;margin:0 0 24px}"
        "td,th{border:1px solid #bbb;padding:4px 8px;white-space:pre-wrap}"
        "th{background:#eee} .highlight{border:3px solid #f00!important}</style>",
        "</head><body>",
    ]
    for sheet in workbook.worksheets:
        parts.append(f"<h2>{escape(sheet.title)}</h2><table>")
        for row in sheet.iter_rows():
            values = [cell.value for cell in row]
            if not any(value not in (None, "") for value in values):
                continue
            parts.append("<tr>")
            for cell in row:
                value = cell.value
                if isinstance(value, (datetime, date)):
                    text = value.strftime("%m/%d/%Y")
                else:
                    text = "" if value is None else str(value)
                borders = cell.border
                highlighted = any(
                    getattr(side, "color", None) is not None
                    and getattr(side.color, "rgb", "") in {"FFFF0000", "00FF0000"}
                    for side in (borders.left, borders.right, borders.top, borders.bottom)
                )
                tag = "th" if cell.row == 1 else "td"
                cls = " class='highlight'" if highlighted else ""
                parts.append(f"<{tag}{cls}>{escape(text)}</{tag}>")
            parts.append("</tr>")
        parts.append("</table>")
    parts.append("</body></html>")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    Path(output_path).write_text("".join(parts), encoding="utf-8")
    workbook.close()
    return output_path


def build_document_annotation_service(
    cache_dir: str,
) -> IExpenseDocumentAnnotationService:
    """Composition root: wire format implementations to the application port."""
    return ExpenseDocumentAnnotationService(
        (
            PdfExpenseDocumentAnnotator(),
            ImageExpenseDocumentAnnotator(
                PdfCheckNumberResolver(),
                CodexCliImageRegionFallbackMatcher(),
            ),
            ExcelExpenseDocumentAnnotator(),
        ),
        cache_dir,
    )
