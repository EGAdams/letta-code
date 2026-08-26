import shutil
from pathlib import Path

import pytest

from document_annotation import (
    AnnotationResult,
    CodexCliImageRegionFallbackMatcher,
    ExcelExpenseDocumentAnnotator,
    ExpenseDocumentAnnotationService,
    ExpenseEvidence,
    ImageExpenseDocumentAnnotator,
    ImageRegionMatch,
    IExpenseDocumentAnnotator,
    IExpenseDocumentAnnotationService,
    IImageRegionFallbackMatcher,
    PdfCheckNumberResolver,
    PdfExpenseDocumentAnnotator,
    render_excel_for_browser,
)

# Most OCR tests here monkeypatch `pytesseract.image_to_data`, so they run
# anywhere. The handful below deliberately do not -- they render a real image
# and read it back through the real binary, which is the only way to catch a
# regression in what we ask tesseract for. That binary is installed on the live
# box (DESKTOP-2OBSQMC) and usually not on a dev box, so skip rather than fail:
# a red suite that means "you are not the live box" trains people to ignore red.
requires_tesseract = pytest.mark.skipif(
    shutil.which('tesseract') is None,
    reason='needs the real tesseract binary; installed on the live box')


class FakeAnnotator(IExpenseDocumentAnnotator):
    def __init__(self):
        self.calls = 0

    def supports(self, source_path):
        return source_path.endswith(".fake")

    def annotate(self, source_path, output_path, evidence):
        self.calls += 1
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        Path(output_path).write_bytes(b"highlighted")
        return AnnotationResult(output_path, True, page=2)


def evidence():
    return ExpenseEvidence(
        expense_id=42,
        expense_date="2025-01-22",
        amount="18.40",
        description="MEIJER STORE",
        vendor_key="meijer",
    )


class FakeImageRegionFallbackMatcher(IImageRegionFallbackMatcher):
    def __init__(self, match):
        self.match = match
        self.calls = []

    def find_region(self, source_path, evidence):
        self.calls.append((source_path, evidence))
        return self.match


class UnavailableImageRegionFallbackMatcher(IImageRegionFallbackMatcher):
    def find_region(self, source_path, evidence):
        raise OSError("matching service offline")


def test_image_strategy_uses_confident_fallback_when_established_match_is_absent(
    tmp_path,
    monkeypatch,
):
    from PIL import Image
    import pytesseract

    source = tmp_path / "check.png"
    Image.new("RGB", (200, 120), "white").save(source)
    output = tmp_path / "annotated.png"
    fallback = FakeImageRegionFallbackMatcher(
        ImageRegionMatch(region=(20, 30, 180, 90), confidence=0.99)
    )
    monkeypatch.setattr(
        pytesseract,
        "image_to_data",
        lambda *args, **kwargs: {"text": []},
    )

    result = ImageExpenseDocumentAnnotator(
        fallback_matcher=fallback,
    ).annotate(str(source), str(output), evidence())

    assert fallback.calls == [(str(source), evidence())]
    assert result.highlighted is True
    assert result.path == str(output)
    annotated = Image.open(output)
    red_pixels = [
        (x, y)
        for y in range(annotated.height)
        for x in range(annotated.width)
        if annotated.getpixel((x, y))[0] > 200
        and annotated.getpixel((x, y))[1] < 80
        and annotated.getpixel((x, y))[2] < 80
    ]
    annotated.close()
    assert red_pixels
    assert min(x for x, _y in red_pixels) < 20
    assert max(x for x, _y in red_pixels) > 180
    assert min(y for _x, y in red_pixels) < 30
    assert max(y for _x, y in red_pixels) > 90


@pytest.mark.parametrize(
    "fallback_match",
    [
        None,
        ImageRegionMatch(region=(20, 30, 180, 90), confidence=0.5),
        ImageRegionMatch(region=(-1, 30, 180, 90), confidence=0.99),
    ],
)
def test_image_strategy_fails_closed_for_uncertain_fallback_results(
    tmp_path,
    monkeypatch,
    fallback_match,
):
    from PIL import Image
    import pytesseract

    source = tmp_path / "check.png"
    Image.new("RGB", (200, 120), "white").save(source)
    output = tmp_path / "annotated.png"
    monkeypatch.setattr(
        pytesseract,
        "image_to_data",
        lambda *args, **kwargs: {"text": []},
    )

    result = ImageExpenseDocumentAnnotator(
        fallback_matcher=FakeImageRegionFallbackMatcher(fallback_match),
    ).annotate(str(source), str(output), evidence())

    assert result.highlighted is False
    assert result.path == str(source)
    assert output.exists() is False


def test_image_strategy_fails_closed_when_fallback_service_is_unavailable(
    tmp_path,
    monkeypatch,
):
    from PIL import Image
    import pytesseract

    source = tmp_path / "check.png"
    Image.new("RGB", (200, 120), "white").save(source)
    output = tmp_path / "annotated.png"
    monkeypatch.setattr(
        pytesseract,
        "image_to_data",
        lambda *args, **kwargs: {"text": []},
    )

    result = ImageExpenseDocumentAnnotator(
        fallback_matcher=UnavailableImageRegionFallbackMatcher(),
    ).annotate(str(source), str(output), evidence())

    assert result.highlighted is False
    assert result.path == str(source)
    assert "unavailable" in result.reason.lower()
    assert output.exists() is False


def test_codex_fallback_requests_one_payment_region_in_original_pixels():
    calls = []

    def runner(command, **kwargs):
        calls.append((command, kwargs))

        class Completed:
            returncode = 0
            stdout = (
                '{"confidence":0.98,"regions":['
                '{"left":100,"top":80,"right":900,"bottom":420}]}'
            )
            stderr = ""

        return Completed()

    target = ExpenseEvidence(
        expense_id=2004,
        expense_date="2025-09-08",
        amount="125.00",
        description="John Roark",
        vendor_key="john_roark",
    )

    match = CodexCliImageRegionFallbackMatcher(
        codex_path="/opt/codex",
        runner=runner,
    ).find_region("/receipts/check.jpg", target)

    assert match == ImageRegionMatch(
        region=(100.0, 80.0, 900.0, 420.0),
        confidence=0.98,
    )
    command, kwargs = calls[0]
    assert command[:2] == ["/opt/codex", "exec"]
    assert command[command.index("--image") + 1] == "/receipts/check.jpg"
    assert "125.00" in kwargs["input"]
    assert "John Roark" in kwargs["input"]
    assert "posting" in kwargs["input"].lower()
    assert kwargs["timeout"] > 0


def test_codex_fallback_rejects_multiple_candidate_regions():
    def runner(command, **kwargs):
        class Completed:
            returncode = 0
            stdout = (
                '{"confidence":0.99,"regions":['
                '{"left":10,"top":20,"right":100,"bottom":80},'
                '{"left":10,"top":100,"right":100,"bottom":160}]}'
            )
            stderr = ""

        return Completed()

    matcher = CodexCliImageRegionFallbackMatcher(
        codex_path="/opt/codex",
        runner=runner,
    )

    assert matcher.find_region("/receipts/check.jpg", evidence()) is None


def test_service_programs_to_interface_and_caches_annotated_copy(tmp_path):
    source = tmp_path / "statement.fake"
    source.write_bytes(b"original")
    annotator = FakeAnnotator()
    service: IExpenseDocumentAnnotationService = ExpenseDocumentAnnotationService(
        [annotator], str(tmp_path / "cache")
    )

    first = service.prepare(str(source), evidence())
    second = service.prepare(str(source), evidence())

    assert first.highlighted is True
    assert first.path != str(source)
    assert second.path == first.path
    assert annotator.calls == 1
    assert source.read_bytes() == b"original"


def test_service_fails_closed_for_an_unowned_format(tmp_path):
    source = tmp_path / "legacy.xls"
    source.write_bytes(b"original")
    service = ExpenseDocumentAnnotationService([], str(tmp_path / "cache"))

    result = service.prepare(str(source), evidence())

    assert result.highlighted is False
    assert result.path == str(source)
    assert "no annotation strategy" in result.reason


def test_pdf_strategy_boxes_the_matching_expense_line(tmp_path):
    import fitz

    source = tmp_path / "statement.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_text((72, 100), "01/21/2025 OTHER STORE $18.40")
    page.insert_text((72, 130), "01/22/2025 MEIJER STORE $18.40")
    document.save(source)
    document.close()
    output = tmp_path / "annotated.pdf"

    result = PdfExpenseDocumentAnnotator().annotate(
        str(source), str(output), evidence()
    )

    assert result.highlighted is True
    assert result.page == 1
    assert output.exists()
    annotated = fitz.open(output)
    drawings = annotated[0].get_drawings()
    annotated.close()
    assert any(item.get("color") == (1.0, 0.0, 0.0) for item in drawings)


def test_pdf_strategy_boxes_only_the_matching_side_by_side_check(tmp_path):
    import fitz

    source = tmp_path / "checks.pdf"
    document = fitz.open()
    page = document.new_page()
    for x, number, amount in (
        (50, "11020", "100.00"),
        (245, "11021", "228.00"),
        (430, "11023", "25.00"),
    ):
        page.insert_text((x, 100), number)
        page.insert_text((x + 48, 100), "01/13")
        page.insert_text((x + 100, 100), amount)
    document.save(source)
    document.close()
    output = tmp_path / "annotated.pdf"
    target = ExpenseEvidence(
        expense_id=477,
        expense_date="2025-01-13",
        amount="25.00",
        description="Right to Life",
    )

    result = PdfExpenseDocumentAnnotator().annotate(
        str(source), str(output), target
    )

    assert result.highlighted is True
    annotated = fitz.open(output)
    red_rects = [
        item["rect"]
        for item in annotated[0].get_drawings()
        if item.get("color") == (1.0, 0.0, 0.0)
    ]
    annotated.close()
    assert len(red_rects) == 1
    assert red_rects[0].x0 > 400
    assert red_rects[0].width < 180


def test_check_number_in_description_counts_as_reference_identity():
    from document_annotation import _best_line

    target = ExpenseEvidence(
        expense_id=1996,
        expense_date="2026-01-06",
        amount="303.00",
        description="Donation (Check # 1107)",
        vendor_key="childrens_vision_int_inc",
    )

    region, score, text = _best_line(
        [
            ("8/16/2023 1107 303.00 2,111.00", "target-row"),
            ("4/22/2025 1071 303.00 1,203.50", "other-row"),
        ],
        target,
    )

    assert region == "target-row"
    assert score >= 12
    assert "1107" in text


def test_pdf_reference_resolver_derives_check_number(tmp_path):
    import fitz

    source = tmp_path / "checks.pdf"
    document = fitz.open()
    page = document.new_page()
    page.insert_text((400, 100), "11023")
    page.insert_text((455, 100), "01/13")
    page.insert_text((520, 100), "25.00")
    document.save(source)
    document.close()

    terms = PdfCheckNumberResolver().resolve(
        ExpenseEvidence(
            expense_id=477,
            expense_date="2025-01-13",
            amount="25.00",
            related_document_path=str(source),
        )
    )

    assert terms == ("11023",)


@requires_tesseract
def test_image_strategy_boxes_the_matching_ocr_line(tmp_path):
    from PIL import Image, ImageDraw, ImageFont

    from document_annotation import ImageExpenseDocumentAnnotator

    source = tmp_path / "ledger.png"
    image = Image.new("RGB", (1000, 260), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype(
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 34
    )
    draw.text((30, 35), "01/21/2025 OTHER STORE $18.40", fill="black", font=font)
    draw.text((30, 135), "01/22/2025 MEIJER STORE $18.40", fill="black", font=font)
    image.save(source)
    output = tmp_path / "annotated.png"
    fallback = FakeImageRegionFallbackMatcher(None)

    result = ImageExpenseDocumentAnnotator(fallback_matcher=fallback).annotate(
        str(source), str(output), evidence()
    )

    assert result.highlighted is True
    assert fallback.calls == []
    assert output.exists()
    annotated = Image.open(output)
    # A pixel just outside the OCR text row should belong to the red rectangle.
    red_pixels = sum(
        1
        for red, green, blue in annotated.get_flattened_data()
        if red > 200 and green < 80 and blue < 80
    )
    annotated.close()
    assert red_pixels > 100


def test_bill_summary_candidate_joins_amount_rows_to_the_dated_balance_row():
    from document_annotation import (
        ExpenseEvidence,
        _best_line,
        _bill_summary_windows,
    )

    target = ExpenseEvidence(
        expense_id=1985,
        expense_date="2025-04-10",
        amount="53.06",
        description="DTE",
        vendor_key="dte_04_10_25_53_06",
    )
    # Captured from Tesseract on the 2026-08-02 Freezer scan. The charge and
    # its date are printed on adjacent bill-summary rows, not one OCR line.
    lines = [
        ("Gas Commercial Heating 53.06", (200, 1953, 1306, 1992)),
        ("Total Current Charges 53.06", (180, 2004, 1307, 2041)),
        ("Account Balance as of April 10, 2025 = $216.79", (180, 2062, 1307, 2097)),
    ]

    candidates = lines + _bill_summary_windows(lines)
    region, score, text = _best_line(candidates, target)

    assert score >= 10
    assert "Gas Commercial Heating 53.06" in text
    assert "April 10, 2025" in text
    # Adjacent rows supply date context, but only the actual expense row is boxed.
    assert region == (200.0, 1953.0, 1306.0, 1992.0)


def test_image_strategy_passes_original_scan_path_to_tesseract(tmp_path, monkeypatch):
    from PIL import Image
    import pytesseract

    from document_annotation import ImageExpenseDocumentAnnotator

    source = tmp_path / "dte.jpg"
    Image.new("RGB", (1400, 2200), "white").save(source, dpi=(300, 300))
    output = tmp_path / "annotated.jpg"
    seen = []

    def fake_image_to_data(image, output_type, config):
        seen.append((image, config))
        rows = [
            ("Gas Commercial Heating 53.06", 200, 1953, 1106, 39, 1),
            ("Total Current Charges 53.06", 180, 2004, 1127, 37, 2),
            ("Account Balance as of April 10, 2025 = $216.79", 180, 2062, 1127, 35, 3),
        ]
        data = {key: [] for key in (
            "text", "left", "top", "width", "height",
            "block_num", "par_num", "line_num",
        )}
        for text, left, top, width, height, line_number in rows:
            data["text"].append(text)
            data["left"].append(left)
            data["top"].append(top)
            data["width"].append(width)
            data["height"].append(height)
            data["block_num"].append(1)
            data["par_num"].append(1)
            data["line_num"].append(line_number)
        return data

    monkeypatch.setattr(pytesseract, "image_to_data", fake_image_to_data)
    result = ImageExpenseDocumentAnnotator().annotate(
        str(source),
        str(output),
        ExpenseEvidence(1985, "2025-04-10", "53.06", "DTE", "dte"),
    )

    assert seen == [
        (str(source), "--psm 1"),
        (str(source), "--psm 6"),
    ]
    assert result.highlighted is True


def test_auto_oriented_form_boxes_only_the_amount_word():
    from document_annotation import _best_line, _image_expense_candidates

    # Tesseract PSM 1 auto-orients the upside-down Children's Vision receipt.
    # The date and amount share a visual row, but the drawing bounds must stay
    # on 300.50 rather than stretching across the form to the date field.
    data = {
        "text": ["300.50", "2/14/2025", "Amount", "Date"],
        "left": [1278, 2139, 1056, 2207],
        "top": [1651, 1658, 1703, 1718],
        "width": [150, 222, 241, 104],
        "height": [38, 40, 44, 39],
        "block_num": [16, 15, 16, 15],
        "par_num": [1, 1, 1, 1],
        "line_num": [2, 2, 1, 1],
    }
    target = ExpenseEvidence(
        1987,
        "2025-02-14",
        "300.50",
        "Children's Vision Int. Inc.",
        "childrens_vision_int_inc",
    )

    region, score, text = _best_line(_image_expense_candidates(data), target)

    assert score >= 10
    assert "2/14/2025" in text
    assert region == (1278.0, 1651.0, 1428.0, 1689.0)


def test_ocr_damaged_total_outranks_exact_approval_confirmation():
    from document_annotation import _best_line

    target = ExpenseEvidence(
        561,
        "2025-02-20",
        "33.13",
        "MR BURGER RESTAURANT",
        "mr_burger_restaurant",
    )
    lines = [
        ("TOTAL 33.1", (1019, 634, 1566, 658)),
        ("Uta receipt noise TOTAL 33.1", (1019, 560, 1589, 591)),
        ("Approved USD $33.13", (1009, 1324, 1572, 1354)),
    ]

    region, score, text = _best_line(lines, target)

    assert score >= 7
    assert text == "TOTAL 33.1"
    assert region == (1019, 634, 1566, 658)


def test_illegible_amount_still_matches_on_date_plus_payee():
    from document_annotation import _best_line

    # What tesseract returns for a statement row EG has written across: the
    # date and payee survive, "10.59" comes back as "1.59".
    lines = [
        ("01/22/2025 MEIJER STORE 1.40", "row"),
        ("Statement period 01/01/2025 - 01/31/2025", "header"),
    ]

    region, score, text = _best_line(lines, evidence())

    assert region == "row"
    assert text == "01/22/2025 MEIJER STORE 1.40"
    # Held below the decisive band: identity alone never outranks a real
    # amount match, and never survives a tie.
    assert score < 10


def test_illegible_amount_match_needs_the_date_and_the_payee():
    from document_annotation import _best_line

    # Right payee, wrong date — this is a different month's Meijer charge.
    assert _best_line([("02/22/2025 MEIJER STORE", "row")], evidence())[0] is None
    # Right date, no payee overlap — a bare date is not identity.
    assert _best_line([("01/22/2025 balance forward", "row")], evidence())[0] is None


def test_illegible_amount_does_not_treat_a_bare_date_in_description_as_row_date():
    from document_annotation import _best_line

    target = ExpenseEvidence(
        expense_id=1679,
        expense_date="2025-03-05",
        amount="19.50",
        description="Edited Rosemary's 2/15 sermon",
    )
    lines = [
        ("3/25/2025 Edited Rosemary's 3/5 sermon 15.50", "wrong-row"),
        ("3/5/2025 Edited Rosemary's 2/15 sermon", "right-row"),
    ]

    assert _best_line(lines, target)[0] == "right-row"


def test_wrapped_invoice_row_matches_on_unique_description_when_date_and_amount_are_garbled():
    from document_annotation import _best_line, _wrapped_line_windows

    target = ExpenseEvidence(
        expense_id=1680,
        expense_date="2025-03-12",
        amount="18.00",
        description=(
            "Edited and uploaded the students and Rosemary's messages on "
            "Love Not the World and Loving the Family of God"
        ),
        vendor_key="jacob_menninga",
    )
    # Captured from Tesseract on the real Freezer scan. The target row wraps,
    # while OCR turns both 3/12/2025 and 18.00 into noise.
    spatial_lines = [
        (
            "Travel Edited and uploaded the-students and Rosemary's messages on a",
            (255.0, 1599.0, 2506.0, 1653.0),
        ),
        (
            "apap avzS Love Not the World and Loving the Family of God "
            "<2 saree —_ Sree?",
            (724.0, 1639.0, 2459.0, 1694.0),
        ),
        (
            "3/19/2025 Edited and uploaded Rosemary's sermon on the golden rule "
            "0.65 15.00 9.75",
            (725.0, 1777.0, 2131.0, 1816.0),
        ),
    ]

    candidates = spatial_lines + _wrapped_line_windows(spatial_lines)
    region, score, text = _best_line(candidates, target)

    assert region == (255.0, 1599.0, 2506.0, 1694.0)
    assert score == 9
    assert "students" in text
    assert "Love Not the World" in text


def test_description_only_fallback_rejects_short_or_repeated_descriptions():
    from document_annotation import _best_line

    short = ExpenseEvidence(
        expense_id=1,
        amount="18.00",
        description="Monthly subscription",
    )
    repeated = ExpenseEvidence(
        expense_id=2,
        amount="18.00",
        description=(
            "Edited and uploaded the students and Rosemary's messages on "
            "Love Not the World and Loving the Family of God"
        ),
    )

    assert _best_line([("Monthly subscription", "row")], short)[0] is None
    assert _best_line(
        [
            (repeated.description, "first-row"),
            (repeated.description, "second-row"),
        ],
        repeated,
    )[0] is None


def test_illegible_amount_match_rejects_two_rows_of_the_same_payee_and_date():
    from document_annotation import _best_line

    # Without a readable amount there is nothing left to tell these apart, so
    # the annotator must box neither rather than guess.
    lines = [
        ("01/22/2025 MEIJER STORE 1.40", "first"),
        ("01/22/2025 MEIJER STORE 8.40", "second"),
    ]

    assert _best_line(lines, evidence())[0] is None


def test_unique_amount_boxes_a_total_line_with_no_date_or_payee_on_it():
    from document_annotation import _best_line

    # AT&T-bill / single-item-receipt shape: the "Total due" line carries no
    # date or vendor text, so it can't clear the normal score-7 bar on its
    # own — but $171.17 appears nowhere else in the document, so there is no
    # rival row it could be confused with.
    target = ExpenseEvidence(
        expense_id=1975,
        expense_date="2025-03-01",
        amount="171.17",
        description="AT&T",
        vendor_key="at_t",
    )
    lines = [
        ("Bill date 03/01/25", "header"),
        ("Total due $171.17", "total-row"),
        ("Account charges $15.00", "line-1"),
    ]

    region, score, text = _best_line(lines, target)

    assert region == "total-row"
    assert score < 7


def test_repeated_amount_still_rejects_the_ambiguous_total_line():
    from document_annotation import _best_line

    # Same shape, but the amount is not unique in the document (it also
    # appears as an unrelated line item) — uniqueness can't rescue this one.
    target = ExpenseEvidence(
        expense_id=1975,
        expense_date="2025-03-01",
        amount="15.00",
        description="AT&T",
        vendor_key="at_t",
    )
    lines = [
        ("Bill date 03/01/25", "header"),
        ("Total due $15.00", "total-row"),
        ("Late fee $15.00", "other-row"),
    ]

    assert _best_line(lines, target)[0] is None


def test_repeated_total_restatement_boxes_the_bottom_most_line():
    from document_annotation import _best_line

    # Payment-slip shape (Dermatology Associates freezer scan, 2026-08-02):
    # a single total is printed twice — a labeled field, then a footer
    # confirmation — with no date or payee on either line. Both disjoint
    # regions carry the word "total", so this is the receipt restating its
    # one figure, not two coincidentally-equal line items; the bottom-most
    # restatement (closest to the signature/copy line) wins.
    target = ExpenseEvidence(
        expense_id=1980,
        expense_date="2025-04-08",
        amount="150.00",
        description="Dermatology Associates of West Michigan",
        vendor_key="dermatology_associates_of_west_michigan_dawm",
    )
    lines = [
        ("Total Amount 150.00", (30, 400, 300, 430)),
        ("Total $150.00", (30, 500, 260, 530)),
    ]

    region, score, text = _best_line(lines, target)

    assert region == (30, 500, 260, 530)
    assert score < 7



@requires_tesseract
def test_image_strategy_boxes_the_amount_column_when_ocr_loses_it(tmp_path):
    from PIL import Image, ImageDraw, ImageFont

    from document_annotation import ImageExpenseDocumentAnnotator

    source = tmp_path / "statement.png"
    image = Image.new("RGB", (1400, 260), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype(
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 34
    )
    draw.text((30, 35), "01/21/2025 OTHER STORE", fill="black", font=font)
    draw.text((1150, 35), "22.10", fill="black", font=font)
    draw.text((30, 135), "01/22/2025 MEIJER STORE", fill="black", font=font)
    # The amount column of the row we want, rendered unreadable the way EG's
    # pen does it on a real scan.
    draw.text((1150, 135), "18.40", fill="black", font=font)
    draw.line((1120, 175, 1360, 130), fill="black", width=9)
    draw.line((1120, 130, 1360, 175), fill="black", width=9)
    image.save(source)
    output = tmp_path / "annotated.png"

    result = ImageExpenseDocumentAnnotator().annotate(
        str(source), str(output), evidence()
    )

    assert result.highlighted is True
    annotated = Image.open(output)
    pixels = annotated.load()
    # The box must reach across the amount column, not stop at the payee text.
    red_columns = {
        x
        for x in range(annotated.width)
        for y in range(annotated.height)
        if pixels[x, y][0] > 200
        and pixels[x, y][1] < 80
        and pixels[x, y][2] < 80
    }
    annotated.close()
    assert red_columns
    assert max(red_columns) > 1200


def test_excel_strategy_boxes_only_the_matching_row(tmp_path):
    from openpyxl import Workbook, load_workbook

    source = tmp_path / "ledger.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Date", "Description", "Amount"])
    sheet.append(["01/21/2025", "OTHER STORE", 18.40])
    sheet.append(["01/22/2025", "MEIJER STORE", 18.40])
    workbook.save(source)
    output = tmp_path / "annotated.xlsx"

    result = ExcelExpenseDocumentAnnotator().annotate(
        str(source), str(output), evidence()
    )

    assert result.highlighted is True
    highlighted = load_workbook(output)
    assert highlighted.active["A3"].border.left.color.rgb == "FFFF0000"
    assert highlighted.active["B3"].border.top.color.rgb == "FFFF0000"
    assert highlighted.active["C3"].border.right.color.rgb == "FFFF0000"
    assert highlighted.active["A2"].border.left.style is None
    highlighted.close()


def test_render_excel_for_browser_preserves_red_highlight(tmp_path):
    from openpyxl import Workbook

    source = tmp_path / 'annotated.xlsx'
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(['Date', 'Description', 'Amount'])
    sheet.append(['01/21/2025', 'TRINITY HEALTH CORP-LIVONIA MI', 225.00])
    from openpyxl.styles import Border, Side
    red = Side(style='thick', color='FFFF0000')
    sheet['B2'].border = Border(left=red, right=red, top=red, bottom=red)
    workbook.save(source)
    output = tmp_path / 'view.html'

    render_excel_for_browser(str(source), str(output))

    html = output.read_text()
    assert 'TRINITY HEALTH CORP-LIVONIA MI' in html
    assert 'class=\'highlight\'' in html



def test_card_statement_row_identifies_by_bare_month_day_and_payee_head():
    """Card statements print the year once, in the billing-cycle header.

    Every transaction row opens with "MM/DD MM/DD", so requiring a printed
    calendar date made the identity-only path unreachable on scans of them —
    the Choice Privileges freezer scan of 2026-07-29 lost the box on two rows
    whose amount OCR could not read.
    """
    from document_annotation import _line_score

    roku = ExpenseEvidence(
        expense_id=1695,
        expense_date="2025-03-17",
        amount="7.99",
        description="THE ROKU CHANNEL 8162728107 DE",
        vendor_key="the_roku_channel",
    )
    radisson = ExpenseEvidence(
        expense_id=1476,
        expense_date="2025-03-15",
        amount="127.56",
        description="RADISSON HOTELS GRAND GRAND RAPIDS ,MI",
        vendor_key="radisson",
    )

    # OCR read the amount as "BF .99" and split the year off the row.
    assert _line_score(
        "03/17 03/1 7 210001500 15270212Q009MWRO6 THE ROKU CHANNEL "
        "8162728107 DE BF .99 | Lu=",
        roku,
    ) >= 7
    # OCR truncated the payee and dropped the amount column entirely.
    assert _line_score(
        "03/15 03/15 920001300 85369432BAS8HP1E3 RADISSON HO", radisson
    ) >= 7


def test_bare_month_day_needs_the_payee_head_not_its_city():
    from document_annotation import _line_score

    radisson = ExpenseEvidence(
        expense_id=1476,
        expense_date="2025-03-15",
        amount="127.56",
        description="RADISSON HOTELS GRAND GRAND RAPIDS ,MI",
        vendor_key="radisson",
    )

    # A different merchant in the same town, on the same day: "GRAND RAPIDS MI"
    # is shared by every row printed from that city and identifies nothing.
    assert _line_score(
        "03/15 03/15 000 SOME OTHER PLACE GRAND RAPIDS MI", radisson
    ) < 0
    # A bare M/D inside prose is not a transaction row either.
    assert _line_score(
        "your payment must be received by 03/15 to avoid a late fee", radisson
    ) < 0


@requires_tesseract
def test_image_strategy_boxes_a_card_statement_row_without_a_printed_year(
    tmp_path,
):
    from PIL import Image, ImageDraw, ImageFont

    from document_annotation import ImageExpenseDocumentAnnotator

    source = tmp_path / "card_statement.png"
    image = Image.new("RGB", (1500, 320), "white")
    draw = ImageDraw.Draw(image)
    font = ImageFont.truetype(
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 32
    )
    draw.text((30, 30), "Billing Cycle 02/20/2025 to 03/21/2025",
              fill="black", font=font)
    draw.text((30, 120), "03/15 03/15 920001300 RADISSON HOTELS GRAND",
              fill="black", font=font)
    draw.text((30, 210), "03/17 03/17 210001500 THE ROKU CHANNEL 8162728107",
              fill="black", font=font)
    # The amount column both rows lost to EG's pen.
    draw.text((1290, 210), "7.99", fill="black", font=font)
    draw.line((1270, 250, 1460, 205), fill="black", width=9)
    draw.line((1270, 205, 1460, 250), fill="black", width=9)
    image.save(source)
    output = tmp_path / "annotated.png"

    result = ImageExpenseDocumentAnnotator().annotate(
        str(source),
        str(output),
        ExpenseEvidence(
            expense_id=1695,
            expense_date="2025-03-17",
            amount="7.99",
            description="THE ROKU CHANNEL 8162728107 DE",
            vendor_key="the_roku_channel",
        ),
    )

    assert result.highlighted is True
    annotated = Image.open(output)
    pixels = annotated.load()
    red_rows = {
        y
        for y in range(annotated.height)
        for x in range(annotated.width)
        if pixels[x, y][0] > 200
        and pixels[x, y][1] < 80
        and pixels[x, y][2] < 80
    }
    annotated.close()
    assert red_rows
    # The box must sit on the ROKU row, not the RADISSON row above it.
    assert min(red_rows) > 150


@requires_tesseract
def test_childrens_vision_check_1107_expense_1996_annotates_correctly(tmp_path):
    """Expense 1996 (Children's Vision Int. Inc., check #1107) should box correctly.

    This test verifies the specific user-reported case: a donation check receipt
    should be annotated with a red box around the amount and date when viewed.
    """
    from PIL import Image, ImageDraw, ImageFont
    from document_annotation import ImageExpenseDocumentAnnotator

    # Simulate the Children's Vision donation receipt layout
    source = tmp_path / "childrens_vision_check_1107.jpg"
    image = Image.new("RGB", (2480, 3508), "white")  # Letter size at 300 DPI
    draw = ImageDraw.Draw(image)

    # Try to load a truetype font, fall back to default if unavailable
    try:
        font = ImageFont.truetype(
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 48
        )
    except (OSError, IOError):
        # Font file not found, use Pillow's default bitmap font
        # Scale it up by drawing multiple times slightly offset
        font = ImageFont.load_default()

    # Receipt header
    draw.text((100, 100), "Children's Vision Int. Inc.", fill="black", font=font)
    draw.text((100, 200), "Donation Receipt", fill="black", font=font)

    # Critical fields that should be matched
    draw.text((100, 800), "Date: 01/06/2026", fill="black", font=font)
    draw.text((100, 900), "Check #: 1107", fill="black", font=font)
    draw.text((100, 1000), "Amount: $3,047.00", fill="black", font=font)
    draw.text((100, 1100), "Thank you for your donation", fill="black", font=font)

    image.save(source)
    output = tmp_path / "annotated.jpg"

    # Evidence matching expense_id 1996
    evidence = ExpenseEvidence(
        expense_id=1996,
        expense_date="2026-01-06",
        amount="3047.00",
        description="Children's Vision Int. Inc. — Donation (Check # 1107)",
        vendor_key="childrens_vision_int_inc",
    )

    result = ImageExpenseDocumentAnnotator().annotate(
        str(source), str(output), evidence
    )

    assert result.highlighted is True, f"Expected highlighting but got: {result.reason}"
    assert output.exists()

    # Verify the red box is present
    annotated = Image.open(output)
    pixels = annotated.load()
    red_pixels = sum(
        1
        for y in range(annotated.height)
        for x in range(annotated.width)
        if pixels[x, y][0] > 200
        and pixels[x, y][1] < 80
        and pixels[x, y][2] < 80
    )
    annotated.close()
    assert red_pixels > 100, "Red box should be visible in the annotated image"


def test_priority_health_eob_boxes_provider_billed_not_intro_text():
    """Priority Health EOBs: intro text has date+description but wrong region.

    The EOB intro says "ROSEMARY L. BARNES had a service visit on January 28, 2025"
    which matches the date and description. If _bill_summary_windows creates a
    candidate joining the intro, amount summary, and provider-billed lines, we must
    box the Provider billed line (where the expense actually appears), NOT the intro.
    """
    from document_annotation import _best_line, _bill_summary_windows

    target = ExpenseEvidence(
        expense_id=1988,
        expense_date="2025-01-28",
        amount="240.00",
        description="Priority Health - OLENZER, EMILY K",
        vendor_key="priority_health",
    )

    # Realistic scenario: intro text is close enough to the amount lines that
    # _bill_summary_windows creates a composite candidate
    lines = [
        ("ROSEMARY L. BARNES had a service visit on January 28, 2025 with OLENZER, EMILY K", (30, 550, 600, 570)),
        ("Your claim summary", (30, 580, 600, 595)),
        ("Provider billed $240.00", (30, 597, 400, 615)),
    ]

    candidates = lines + _bill_summary_windows(lines)
    region, score, text = _best_line(candidates, target)

    assert region is not None, f"Should find a match. Candidates: {len(candidates)}"
    assert score >= 10, f"Should have high confidence with date+amount, got score {score}"

    # The box MUST be around "Provider billed $240.00" (y=597), NOT the intro text (y=550)
    assert region[1] >= 590 and region[1] <= 620, (
        f"BUG: Boxed intro text (y={region[1]}) instead of Provider billed line (y~597). "
        f"Matched text: {text}"
    )
    assert "240.00" in text, f"Matched text should contain the target amount, got: {text}"


def _diners_freezer_ocr_data():
    rows = [
        ("05/30/2025 DIN-2025-0587 ADINUAL FEE Diners Club x-0587 (9500 )", 150),
        ("06/15/2025 DIN-2025-0588 OTHER CHARGE 45.00", 250),
    ]
    data = {
        key: []
        for key in ("text", "block_num", "par_num", "line_num", "left", "top", "width", "height")
    }
    for line_number, (line, top) in enumerate(rows, 1):
        left = 50
        for token in line.split():
            width = max(12, len(token) * 18)
            data["text"].append(token)
            data["block_num"].append(1)
            data["par_num"].append(1)
            data["line_num"].append(line_number)
            data["left"].append(left)
            data["top"].append(top)
            data["width"].append(width)
            data["height"].append(32)
            left += width + 8
    return data


def test_last_freezer_scan_boxes_annual_fee_row_2000(tmp_path, monkeypatch):
    """Regression: Diners Club ANNUAL FEE row 2000 must be boxed in scanned JPG.

    This test verifies that ImageExpenseDocumentAnnotator can highlight the
    expense line from the real Freezer scan JPEG, even when the row is complex.
    """
    from PIL import Image, ImageDraw, ImageFont
    from document_annotation import ImageExpenseDocumentAnnotator

    source = tmp_path / "diners_club_0587_april_22__may_30.jpg"
    image = Image.new("RGB", (1400, 500), "white")
    draw = ImageDraw.Draw(image)

    try:
        font = ImageFont.truetype(
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 32
        )
    except (OSError, IOError):
        font = ImageFont.load_default()

    # Simulate relevant lines from the Diners statement
    draw.text((50, 50), "Transaction Date    Posted Date    Reference Number    Description", fill="black", font=font)
    draw.text((50, 150), "05/30/2025    05/30/2025    DIN-2025-0587    ANNUAL FEE Diners Club | x-0587    95.00", fill="black", font=font)
    draw.text((50, 250), "06/15/2025    06/15/2025    DIN-2025-0588    OTHER CHARGE    45.00", fill="black", font=font)

    image.save(source)
    output = tmp_path / "annotated.jpg"

    import pytesseract

    monkeypatch.setattr(
        pytesseract,
        "image_to_data",
        lambda *args, **kwargs: _diners_freezer_ocr_data(),
    )

    evidence = ExpenseEvidence(
        expense_id=2000,
        expense_date="2025-05-30",
        amount="95.00",
        description="ANNUAL FEE Diners Club | x-0587",
        vendor_key="diners_club_0587",
    )

    result = ImageExpenseDocumentAnnotator().annotate(
        str(source), str(output), evidence
    )

    assert result.highlighted is True
    assert output.exists()

    # Verify output bytes differ from source (highlighting applied)
    source_bytes = source.read_bytes()
    output_bytes = output.read_bytes()
    assert source_bytes != output_bytes


def test_best_line_joins_split_diners_annual_fee_description():
    """The real scan splits the fee label and card identifier across rows."""
    from document_annotation import _best_line

    evidence = ExpenseEvidence(
        expense_id=2000,
        expense_date="2025-05-30",
        amount="95.00",
        description="ANNUAL FEE Diners Club | x-0587",
        vendor_key="diners_club_0587",
    )
    region, score, text = _best_line(
        [
            ("3 Months a A)INUAL FEE", (220, 1352, 1147, 1423)),
            ("Diners Club |-x-0587", (893, 1456, 1226, 1491)),
            ("OTHER CHARGE", (893, 1560, 1226, 1595)),
        ],
        evidence,
    )

    assert region == (220, 1352, 1226, 1491)
    assert score >= 7
    assert "Diners Club" in text


def test_row_expansion_boxes_the_complete_walgreens_statement_row():
    """Walgreens 2003 must show date, merchant, and amount in one red box."""
    from document_annotation import _expand_selected_row_region

    evidence = ExpenseEvidence(
        expense_id=2003,
        expense_date="2025-01-17",
        amount="141.76",
        description="WALGREENS #15466 GRAND RAPIDS MI",
        vendor_key="walgreens",
    )
    # This is the shape produced when the PDF/OCR extractor enumerates the
    # three cells separately. The date/amount candidate won, but the old
    # drawing boundary stopped at that candidate's date fragment.
    lines = [
        ("01/17/2025", (100.0, 240.0, 185.0, 262.0)),
        ("WALGREENS #15466 GRAND RAPIDS MI", (205.0, 240.0, 650.0, 262.0)),
        ("141.76", (1120.0, 240.0, 1195.0, 262.0)),
        # A second Walgreens charge on another row must not be pulled in.
        (
            "01/23/2025 WALGREENS #15466 GRAND RAPIDS MI 9.85",
            (100.0, 280.0, 1195.0, 302.0),
        ),
    ]

    region, text = _expand_selected_row_region(lines, lines[0][1], evidence)

    assert region == (100.0, 240.0, 1195.0, 262.0)
    assert "WALGREENS" in text
    assert "141.76" in text
    assert "9.85" not in text


def test_image_strategy_boxes_the_complete_walgreens_statement_row(tmp_path, monkeypatch):
    """The image strategy must apply generic row expansion, not just PDF."""
    from PIL import Image
    import document_annotation
    from document_annotation import ImageExpenseDocumentAnnotator

    source = tmp_path / "walgreens.png"
    output = tmp_path / "annotated.png"
    Image.new("RGB", (1300, 420), "white").save(source)

    lines = [
        ("01/17/2025", (100.0, 120.0, 185.0, 142.0)),
        ("WALGREENS #15466 GRAND RAPIDS MI", (205.0, 120.0, 650.0, 142.0)),
        ("141.76", (1120.0, 120.0, 1195.0, 142.0)),
        ("01/23/2025 WALGREENS #15466 GRAND RAPIDS MI 9.85", (100.0, 180.0, 1195.0, 202.0)),
    ]

    import pytesseract

    monkeypatch.setattr(
        pytesseract,
        "image_to_data",
        lambda *args, **kwargs: {},
    )
    monkeypatch.setattr(
        document_annotation,
        "_image_expense_candidates",
        lambda _data: lines,
    )

    evidence = ExpenseEvidence(
        expense_id=2003,
        expense_date="2025-01-17",
        amount="141.76",
        description="WALGREENS #15466 GRAND RAPIDS MI",
        vendor_key="walgreens",
    )
    result = ImageExpenseDocumentAnnotator().annotate(
        str(source), str(output), evidence
    )

    assert result.highlighted is True
    annotated = Image.open(output).convert("RGB")
    # The expanded red rectangle includes the date and amount cells. Its
    # bottom is below the first row but above the neighboring row.
    assert annotated.getpixel((100, 112))[0] > 200
    assert annotated.getpixel((1195, 112))[0] > 200
    assert annotated.getpixel((100, 155))[:3] == (255, 255, 255)


def test_amount_in_description_with_ambiguous_rows_fails_closed():
    """When description contains the amount and multiple rows match, fail closed.

    Regression test for the bug where "Affordable Store 61.80" matched multiple
    rows with the same date and amount, but the description tokens ("store")
    were too generic to disambiguate them.
    """
    from document_annotation import _best_line

    # Case 1: Amount appears in description, multiple rows have same date+amount
    evidence1 = ExpenseEvidence(
        expense_id=9001,
        expense_date="2026-08-05",
        amount="61.80",
        description="Affordable Store 61.80",
        vendor_key="affordable_store",
    )

    # All three rows match on date+amount, but only "store" token overlaps
    lines1 = [
        ("08/05/2026 First Store 61.80", (100, 200, 800, 230)),
        ("08/05/2026 Second Store 61.80", (100, 250, 800, 280)),
        ("08/05/2026 Third Store 61.80", (100, 300, 800, 330)),
    ]

    region, score, text = _best_line(lines1, evidence1)
    # Should fail closed - can't distinguish which "Store" is the right one
    assert region is None, (
        f"Should fail closed on ambiguous match, but boxed: {text!r}"
    )

    # Case 2: Distinctive description disambiguates despite same date+amount
    evidence2 = ExpenseEvidence(
        expense_id=9002,
        expense_date="2026-08-05",
        amount="61.80",
        description="Second Store",  # Distinctive token "Second"
        vendor_key="second_store",
    )

    region, score, text = _best_line(lines1, evidence2)
    # Should succeed - "Second" uniquely identifies the middle row
    assert region is not None, "Should box the distinctive match"
    assert text == "08/05/2026 Second Store 61.80"


def test_typo_in_description_with_correct_ocr_still_matches():
    """Typo in description ('Annuall' vs 'ANNUAL') shouldn't prevent matching.

    Regression test for "Annuall Fee 95.00" which had a typo but should still
    match the correctly OCR'd "ANNUAL FEE" line via the OCR alias system and
    other matching description tokens.
    """
    from document_annotation import _best_line

    evidence = ExpenseEvidence(
        expense_id=9003,
        expense_date="2025-05-30",
        amount="95.00",
        description="Annuall Fee Diners Club",  # Typo: "Annuall" not "Annual"
        vendor_key="diners_club",
    )

    lines = [
        ("05/30/2025 ANNUAL FEE Diners Club 95.00", (100, 200, 800, 230)),
        ("06/15/2025 OTHER CHARGE 45.00", (100, 250, 800, 280)),
    ]

    region, score, text = _best_line(lines, evidence)

    # Should successfully box the ANNUAL FEE line despite the typo, because:
    # 1. OCR aliases map common misspellings like "anual" -> "annual"
    # 2. "Diners" and "Club" tokens also match
    # 3. Date and amount match
    assert region is not None, f"Should box despite typo, got: {region}"
    assert text == "05/30/2025 ANNUAL FEE Diners Club 95.00"
    assert score >= 10, f"Should have high confidence score, got: {score}"


def test_repeated_amount_picks_cleaner_candidate():
    """When same row appears twice with/without repeated amount, box the cleaner one.

    Regression test for the bug where OCR produces two very similar candidates
    for the same transaction - one with the amount appearing once, and one with
    it repeated (e.g., appearing in multiple columns or OCR fragments). The
    tie-breaking logic should recognize this scenario and prefer the cleaner
    candidate rather than failing closed.
    """
    from document_annotation import _best_line

    evidence = ExpenseEvidence(
        expense_id=9004,
        expense_date="2026-08-05",
        amount="61.80",
        description="Affordable Store",
        vendor_key="affordable_store",
    )

    # Two candidates: same transaction, but one has amount repeated
    lines = [
        ("08/05/2026 Affordable Store $61.80 $61.80", (100, 200, 800, 230)),
        ("08/05/2026 Affordable Store $61.80", (100, 250, 800, 280)),
    ]

    region, score, text = _best_line(lines, evidence)

    # Should box the cleaner version (without repetition)
    assert region is not None, "Should box the cleaner candidate, not fail closed"
    assert text == "08/05/2026 Affordable Store $61.80"
    assert region == (100, 250, 800, 280), f"Should pick the non-repeated version, got {region}"


def test_neighboring_amount_in_composite_window_cannot_steal_row():
    """A wrapped OCR window containing a neighboring charge must lose."""
    from document_annotation import _best_line

    evidence = ExpenseEvidence(
        expense_id=9005,
        expense_date="2025-04-24",
        amount="61.80",
        description="AFFORDABLE I STORE ... Diners Club | x-0587",
        vendor_key="affordable_i_store",
    )
    correct = (
        "April 24, 2025 AFFORDABLE I STORE ... Other Expenses -$61.80",
        (100, 300, 900, 340),
    )
    # This is the shape produced when an evidence-aware window joins the
    # neighboring $264.99 row to the target $61.80 row.
    composite = (
        "AFFORDABLE I STORE ... $264.99 April 24, 2025 "
        "AFFORDABLE I STORE ... -$61.80",
        (100, 220, 1100, 340),
    )

    region, _score, text = _best_line([composite, correct], evidence)

    assert region == correct[1]
    assert text == correct[0]
